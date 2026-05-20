from __future__ import annotations

import os
import sys
from datetime import date, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

GCC_COUNTRIES = {"Saudi Arabia", "United Arab Emirates", "Kuwait", "Qatar", "Oman", "Bahrain"}
EXCLUDE_EN_NON_GCC = {"Singapore", "Malaysia"}
ALL_DESK_COUNTRIES = {"Mauritius", "Brunei Darussalam"}


def extract_desk_code(desk_name):
    if pd.isna(desk_name):
        return ""
    parts = str(desk_name).split("-")
    return parts[1].strip() if len(parts) >= 2 else str(desk_name).strip()


def load_data(filepath):
    df = pd.read_excel(filepath, header=None, skiprows=3)
    df.columns = range(df.shape[1])
    df = df[[1, 2, 3, 5, 6, 8]].copy()
    df.columns = ["Desk", "Campaign", "Country", "Reg", "FTD", "LateFTD"]
    df = df[~(df["Reg"].isna() & df["FTD"].isna())]
    df["Reg"] = pd.to_numeric(df["Reg"], errors="coerce").fillna(0)
    df["FTD"] = pd.to_numeric(df["FTD"], errors="coerce").fillna(0)
    df["LateFTD"] = pd.to_numeric(df["LateFTD"], errors="coerce").fillna(0)
    df["Desk"] = df["Desk"].apply(extract_desk_code)
    df["Country"] = df["Country"].fillna("").astype(str).str.strip()
    df["Campaign"] = df["Campaign"].fillna("").astype(str).str.strip()
    return df


def cr(ftd, reg):
    return ftd / reg if reg else 0


DARK_BLUE = "1F3864"
MID_BLUE = "2E75B6"
LIGHT_BLUE = "BDD7EE"
ALT_BLUE = "DDEBF7"
WHITE = "FFFFFF"


def fill(hex_):
    return PatternFill("solid", start_color=hex_, fgColor=hex_)


thin = Side(style="thin", color="9DC3E6")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def sc(ws, row, col, value="", bold=False, bg=WHITE, fg="000000", align="left", numfmt=None, merge_to_col=None):
    cell = ws.cell(row, col, value)
    cell.font = Font(
        name="Arial",
        bold=bold,
        color="FFFFFF" if bg in (DARK_BLUE, MID_BLUE) else fg,
        size=10,
    )
    cell.fill = fill(bg)
    cell.border = BORDER
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if numfmt:
        cell.number_format = numfmt
    if merge_to_col:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_to_col)
    return cell


class ColWidths:
    def __init__(self):
        self.widths = {}

    def feed(self, col, text):
        size = len(str(text)) if text is not None else 0
        self.widths[col] = max(self.widths.get(col, 8), size + 3)

    def apply(self, ws):
        for col, width in self.widths.items():
            ws.column_dimensions[get_column_letter(col)].width = min(width, 40)


def build_p1(df):
    return (
        df[df["Country"] != ""]
        .groupby("Country")
        .agg(Reg=("Reg", "sum"), FTD=("FTD", "sum"))
        .reset_index()
        .sort_values("Country")
    )


def build_p2(df):
    return (
        df[(df["Desk"] != "") & (df["Country"] != "")]
        .groupby(["Desk", "Country"])
        .agg(Reg=("Reg", "sum"), FTD=("FTD", "sum"))
        .reset_index()
        .sort_values(["Desk", "Country"])
    )


def build_p3(df):
    sub = df[(df["Desk"] != "") & (df["Country"] != "") & (df["Campaign"] != "") & (df["LateFTD"] > 0)]
    return (
        sub.groupby(["Desk", "Country", "Campaign"])
        .agg(LateFTD=("LateFTD", "sum"))
        .reset_index()
        .sort_values(["Desk", "Country", "Campaign"])
    )


def build_p4(df):
    sub = df[df["Country"].isin(GCC_COUNTRIES) & df["Campaign"].str.startswith("M-Inhousemedia")]
    return sub.groupby("Country").agg(Reg=("Reg", "sum"), FTD=("FTD", "sum")).reset_index().sort_values("Country")


def build_p5(df):
    sub = df[df["Country"].isin(GCC_COUNTRIES) & (df["Desk"] == "AR") & ~df["Campaign"].str.startswith("M-Inhousemedia")]
    return sub.groupby("Country").agg(Reg=("Reg", "sum"), FTD=("FTD", "sum")).reset_index().sort_values("Country")


def build_p6(df):
    sub = df[df["Country"].isin(GCC_COUNTRIES) & (df["Desk"] == "EN") & ~df["Campaign"].str.startswith("M-Inhousemedia")]
    return sub.groupby("Country").agg(Reg=("Reg", "sum"), FTD=("FTD", "sum")).reset_index().sort_values("Country")


def build_p7(df):
    en_sub = df[
        (df["Desk"] == "EN")
        & ~df["Country"].isin(GCC_COUNTRIES)
        & ~df["Country"].isin(EXCLUDE_EN_NON_GCC)
        & ~df["Country"].isin(ALL_DESK_COUNTRIES)
        & (df["Country"] != "")
    ]
    all_sub = df[df["Country"].isin(ALL_DESK_COUNTRIES)]
    sub = pd.concat([en_sub, all_sub])
    return sub.groupby("Country").agg(Reg=("Reg", "sum"), FTD=("FTD", "sum")).reset_index().sort_values("Country")


def write_simple(ws, cw, start_row, start_col, title, grouped, ncols=4):
    col = start_col
    row = start_row

    sc(ws, row, col, title, bold=True, bg=DARK_BLUE, merge_to_col=col + ncols - 1, align="center")
    for c in range(col + 1, col + ncols):
        ws.cell(row, c).border = BORDER
    cw.feed(col, title)
    row += 1

    for ci, header in enumerate(["Country", "Regs", "FTD", "CR"], col):
        sc(ws, row, ci, header, bold=True, bg=MID_BLUE, align="center")
        cw.feed(ci, header)
    row += 1

    alt = False
    for _, grouped_row in grouped.iterrows():
        bg = ALT_BLUE if alt else WHITE
        alt = not alt
        regs = int(grouped_row.Reg)
        ftd = int(grouped_row.FTD)
        sc(ws, row, col, grouped_row.Country, bg=bg, align="left")
        sc(ws, row, col + 1, regs, bg=bg, align="center", numfmt="#,##0")
        sc(ws, row, col + 2, ftd, bg=bg, align="center", numfmt="#,##0")
        sc(ws, row, col + 3, cr(ftd, regs), bg=bg, align="center", numfmt="0%")
        cw.feed(col, grouped_row.Country)
        cw.feed(col + 1, regs)
        cw.feed(col + 2, ftd)
        cw.feed(col + 3, "00%")
        row += 1

    total_reg = int(grouped.Reg.sum())
    total_ftd = int(grouped.FTD.sum())
    sc(ws, row, col, "Grand Total", bold=True, bg=DARK_BLUE, align="left")
    sc(ws, row, col + 1, total_reg, bold=True, bg=DARK_BLUE, align="center", numfmt="#,##0")
    sc(ws, row, col + 2, total_ftd, bold=True, bg=DARK_BLUE, align="center", numfmt="#,##0")
    sc(ws, row, col + 3, cr(total_ftd, total_reg), bold=True, bg=DARK_BLUE, align="center", numfmt="0%")
    row += 1
    return row


def write_desk_country(ws, cw, start_row, start_col, grouped):
    col = start_col
    row = start_row

    sc(ws, row, col, "Desk + Country", bold=True, bg=DARK_BLUE, merge_to_col=col + 4, align="center")
    for c in range(col + 1, col + 5):
        ws.cell(row, c).border = BORDER
    row += 1

    for ci, header in enumerate(["Desk", "Country", "Regs", "FTD", "CR"], col):
        sc(ws, row, ci, header, bold=True, bg=MID_BLUE, align="center")
        cw.feed(ci, header)
    row += 1

    alt = False
    for desk, desk_group in grouped.groupby("Desk", sort=True):
        first = True
        for _, grouped_row in desk_group.iterrows():
            bg = ALT_BLUE if alt else WHITE
            alt = not alt
            regs = int(grouped_row.Reg)
            ftd = int(grouped_row.FTD)
            desk_val = desk if first else ""
            first = False
            sc(ws, row, col, desk_val, bg=bg, align="left")
            sc(ws, row, col + 1, grouped_row.Country, bg=bg, align="left")
            sc(ws, row, col + 2, regs, bg=bg, align="center", numfmt="#,##0")
            sc(ws, row, col + 3, ftd, bg=bg, align="center", numfmt="#,##0")
            sc(ws, row, col + 4, cr(ftd, regs), bg=bg, align="center", numfmt="0%")
            cw.feed(col, desk_val)
            cw.feed(col + 1, grouped_row.Country)
            row += 1

        desk_total_reg = int(desk_group.Reg.sum())
        desk_total_ftd = int(desk_group.FTD.sum())
        sc(ws, row, col, f"{desk} Total", bold=True, bg=LIGHT_BLUE, align="left")
        sc(ws, row, col + 1, "", bold=True, bg=LIGHT_BLUE)
        sc(ws, row, col + 2, desk_total_reg, bold=True, bg=LIGHT_BLUE, align="center", numfmt="#,##0")
        sc(ws, row, col + 3, desk_total_ftd, bold=True, bg=LIGHT_BLUE, align="center", numfmt="#,##0")
        sc(ws, row, col + 4, cr(desk_total_ftd, desk_total_reg), bold=True, bg=LIGHT_BLUE, align="center", numfmt="0%")
        cw.feed(col, f"{desk} Total")
        row += 1

    total_reg = int(grouped.Reg.sum())
    total_ftd = int(grouped.FTD.sum())
    sc(ws, row, col, "Grand Total", bold=True, bg=DARK_BLUE, align="left")
    sc(ws, row, col + 1, "", bold=True, bg=DARK_BLUE)
    sc(ws, row, col + 2, total_reg, bold=True, bg=DARK_BLUE, align="center", numfmt="#,##0")
    sc(ws, row, col + 3, total_ftd, bold=True, bg=DARK_BLUE, align="center", numfmt="#,##0")
    sc(ws, row, col + 4, cr(total_ftd, total_reg), bold=True, bg=DARK_BLUE, align="center", numfmt="0%")
    row += 1
    return row


def write_latefdt(ws, cw, start_row, start_col, grouped):
    col = start_col
    row = start_row

    sc(ws, row, col, "LATE FTD", bold=True, bg=DARK_BLUE, merge_to_col=col + 3, align="center")
    for c in range(col + 1, col + 4):
        ws.cell(row, c).border = BORDER
    row += 1

    for ci, header in enumerate(["DESK.", "Country", "Campaign", "Late FTD"], col):
        sc(ws, row, ci, header, bold=True, bg=MID_BLUE, align="center")
        cw.feed(ci, header)
    row += 1

    alt = False
    grand_total = 0
    for desk, desk_group in grouped.groupby("Desk", sort=True):
        first_desk = True
        for country, country_group in desk_group.groupby("Country", sort=True):
            country_total = int(country_group.LateFTD.sum())
            grand_total += country_total
            first_country = True
            for _, grouped_row in country_group.iterrows():
                bg = ALT_BLUE if alt else WHITE
                alt = not alt
                desk_val = desk if first_desk else ""
                country_val = country if first_country else ""
                first_desk = False
                first_country = False
                sc(ws, row, col, desk_val, bg=bg, align="left")
                sc(ws, row, col + 1, country_val, bg=bg, align="left")
                sc(ws, row, col + 2, grouped_row.Campaign, bg=bg, align="left")
                sc(ws, row, col + 3, int(grouped_row.LateFTD), bg=bg, align="center", numfmt="#,##0")
                cw.feed(col, desk_val)
                cw.feed(col + 1, country_val)
                cw.feed(col + 2, grouped_row.Campaign)
                row += 1

            sc(ws, row, col, "", bold=True, bg=LIGHT_BLUE)
            sc(ws, row, col + 1, f"{country} Total", bold=True, bg=LIGHT_BLUE, align="left")
            sc(ws, row, col + 2, "", bold=True, bg=LIGHT_BLUE)
            sc(ws, row, col + 3, country_total, bold=True, bg=LIGHT_BLUE, align="center", numfmt="#,##0")
            cw.feed(col + 1, f"{country} Total")
            row += 1

    sc(ws, row, col, "Grand Total", bold=True, bg=DARK_BLUE, merge_to_col=col + 2, align="left")
    for c in range(col + 1, col + 3):
        ws.cell(row, c).border = BORDER
    sc(ws, row, col + 3, grand_total, bold=True, bg=DARK_BLUE, align="center", numfmt="#,##0")
    row += 1
    return row


def process(input_path: Path, output_path: Path) -> None:
    df = load_data(input_path)

    p1 = build_p1(df)
    p2 = build_p2(df)
    p3 = build_p3(df)
    p4 = build_p4(df)
    p5 = build_p5(df)
    p6 = build_p6(df)
    p7 = build_p7(df)

    wb = Workbook()
    ws = wb.active
    ws.title = "Country + desk + late-others"
    ws.sheet_view.showGridLines = False

    cw = ColWidths()

    c1 = 1
    c2 = 6
    c3 = 12
    c4 = 17
    start = 1

    write_simple(ws, cw, start, c1, "All Countries", p1)
    write_desk_country(ws, cw, start, c2, p2)
    write_latefdt(ws, cw, start, c3, p3)

    row = start
    row = write_simple(ws, cw, row, c4, "GCC M-Inhousemedia", p4)
    row += 1
    row = write_simple(ws, cw, row, c4, "Others GCC AR", p5)
    row += 1
    row = write_simple(ws, cw, row, c4, "Others GCC EN", p6)
    row += 1
    write_simple(ws, cw, row, c4, "Others ENG", p7)

    for spacer_col in [5, 11, 16]:
        ws.column_dimensions[get_column_letter(spacer_col)].width = 2

    cw.apply(ws)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def default_output_filename() -> str:
    yesterday = date.today() - timedelta(days=1)
    return f"CR data {yesterday.day}-{yesterday.month}.xlsx"


def main() -> int:
    script_dir = Path(os.path.dirname(os.path.abspath(__file__)))
    input_file = script_dir / "CR.xlsx"

    if not input_file.exists():
        print(f"Error: CR.xlsx not found in {script_dir}")
        return 1

    output_file = script_dir / default_output_filename()
    print(f"Input:  {input_file}")
    print(f"Output: {output_file}")
    process(input_file, output_file)
    print(f"Done! Output saved to: {output_file}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
