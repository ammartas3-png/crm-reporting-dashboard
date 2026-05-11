"""Bulk Country Reports: create CRM country-split workbooks.

Bulk Country Reports shares the CRM/PowerBI merge rules with Report Generator,
then writes:

- a "Main Report" sheet with all rows, a status pivot labelled "M-Inhousemedia",
  and a call-attempts pivot below the data table;
- one sheet per country with formula-linked rows filtered from Main Report and a
  country-labelled status pivot.
"""

from __future__ import annotations

import datetime as _dt
import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from report_generator import (
    CRM_COLUMNS,
    NO_ANSWER_COLORS,
    NO_ANSWER_STATUS_RE,
    NO_COMMENT_STATUSES,
    OUTPUT_COLUMNS,
    POWERBI_COLUMNS,
    STATUS_COLORS,
    comments_for_status,
    extract_comments,
    normalize_header,
    normalize_match_value,
    normalize_status,
)


THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

COLOR_SCALE_RED = "FFF8696B"
COLOR_SCALE_YELLOW = "FFFFEB84"
COLOR_SCALE_GREEN = "FF63BE7B"

PIVOT_START_COL = 6
PIVOT_GAP_ROWS = 3
PIVOT_INTER_GAP = 2
MAIN_REPORT_PIVOT_LABEL = "M-Inhousemedia"
DEFAULT_OUTPUT_FILENAME = "crm_country_report.xlsx"


def worksheet_from_file(path: Path, sheet_name: str | None = None):
    workbook = load_workbook(path, data_only=True)
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"Sheet {sheet_name!r} not found in {path.name}. "
                f"Available sheets: {', '.join(workbook.sheetnames)}"
            )
        return workbook[sheet_name]
    return workbook.active


def header_indexes(
    headers: Iterable[Any],
    required_columns: list[str],
    path: Path,
) -> dict[str, int]:
    normalized_headers = {normalize_header(header): i for i, header in enumerate(headers)}
    missing = [
        column
        for column in required_columns
        if normalize_header(column) not in normalized_headers
    ]
    if missing:
        raise ValueError(f"{path.name} is missing required columns: {', '.join(missing)}")
    return {
        column: normalized_headers[normalize_header(column)]
        for column in required_columns
    }


def _normalize_call_attempts(value: Any) -> int:
    if value is None or value == "":
        return 1
    try:
        number = int(float(str(value).strip()))
    except (TypeError, ValueError):
        return 1
    return number if number >= 1 else 1


def read_powerbi_lookup(
    powerbi_report: Path,
    sheet_name: str | None = None,
) -> dict[tuple[str, str], dict[str, Any]]:
    worksheet = worksheet_from_file(powerbi_report, sheet_name)
    rows = worksheet.iter_rows(values_only=True)

    try:
        headers = next(rows)
    except StopIteration as exc:
        raise ValueError(f"{powerbi_report.name} is empty") from exc

    indexes = header_indexes(headers, POWERBI_COLUMNS, powerbi_report)
    lookup: dict[tuple[str, str], dict[str, Any]] = {}

    for row in rows:
        account_no = row[indexes["Account No"]]
        brand_name = row[indexes["Brand"]]
        account_key = normalize_match_value(account_no)
        brand_key = normalize_match_value(brand_name)

        if not account_key or not brand_key:
            continue

        lookup[(account_key, brand_key)] = {
            "Comments": extract_comments(row[indexes["Last 10 Comments"]]),
            "Call Attempts": _normalize_call_attempts(
                row[indexes["Voip Calls Attempts Cnt"]]
            ),
        }

    return lookup


def read_crm_rows(
    crm_file: Path,
    platform: str,
    powerbi_lookup: dict[tuple[str, str], dict[str, Any]],
    sheet_name: str | None = None,
) -> list[dict[str, Any]]:
    worksheet = worksheet_from_file(crm_file, sheet_name)
    rows = worksheet.iter_rows(values_only=True)

    try:
        headers = next(rows)
    except StopIteration as exc:
        raise ValueError(f"{crm_file.name} is empty") from exc

    indexes = header_indexes(headers, CRM_COLUMNS, crm_file)
    platform_key = normalize_match_value(platform)
    output_rows: list[dict[str, Any]] = []

    for row in rows:
        if all(value is None for value in row):
            continue

        crm_values = {column: row[indexes[column]] for column in CRM_COLUMNS}

        customer_type_norm = normalize_status(crm_values.get("Customer Type", ""))
        if customer_type_norm == "depositor":
            crm_values["Status"] = "Telemarketing"
            output_rows.append(
                {
                    "Platform": platform,
                    **crm_values,
                    "CB": None,
                    "Comments": "",
                    "_comments_yellow": False,
                    "Call Attempts": 1,
                }
            )
            continue

        account_key = normalize_match_value(crm_values["ID"])
        matched = powerbi_lookup.get((account_key, platform_key))
        found_in_powerbi = matched is not None
        matched_comments: list[str] = matched.get("Comments", []) if matched else []

        comment_text, use_yellow = comments_for_status(
            crm_values["Status"],
            matched_comments,
            found_in_powerbi,
        )

        status_norm = normalize_status(crm_values["Status"])
        cb_value = "" if status_norm == "call again" else None

        output_rows.append(
            {
                "Platform": platform,
                **crm_values,
                "CB": cb_value,
                "Comments": comment_text,
                "_comments_yellow": use_yellow,
                "Call Attempts": _normalize_call_attempts(
                    matched.get("Call Attempts") if matched else None
                ),
            }
        )

    return output_rows


def _ff_argb(hex_color: str) -> str:
    value = str(hex_color).upper()
    if len(value) == 6:
        return f"FF{value}"
    return value


def _make_fill(hex_color: str) -> PatternFill:
    color = _ff_argb(hex_color)
    return PatternFill("solid", start_color=color, fgColor=color)


def _make_font(hex_color: str, bold: bool = False) -> Font:
    return Font(color=_ff_argb(hex_color), bold=bold, name="Arial")


def _apply_all_borders(worksheet, total_rows: int, total_cols: int) -> None:
    for row in worksheet.iter_rows(
        min_row=1,
        max_row=total_rows,
        min_col=1,
        max_col=total_cols,
    ):
        for cell in row:
            cell.border = THIN_BORDER


def _color_scale_rule() -> ColorScaleRule:
    return ColorScaleRule(
        start_type="min",
        start_color=COLOR_SCALE_RED,
        mid_type="percentile",
        mid_value=50,
        mid_color=COLOR_SCALE_YELLOW,
        end_type="max",
        end_color=COLOR_SCALE_GREEN,
    )


_INTEGER_RE = re.compile(r"-?\d+")


def _id_literal_for_formula(id_value: Any) -> str:
    if isinstance(id_value, bool):
        return f'"{id_value}"'
    if isinstance(id_value, int):
        return str(id_value)
    if isinstance(id_value, float):
        return str(int(id_value)) if id_value.is_integer() else repr(id_value)
    value = str(id_value).strip()
    if _INTEGER_RE.fullmatch(value):
        return value
    escaped = value.replace('"', '""')
    return f'"{escaped}"'


def _str_literal_for_formula(value: Any) -> str:
    text = "" if value is None else str(value)
    return f'"{text.replace(chr(34), chr(34) * 2)}"'


def _country_cell_formula(
    id_lit: str,
    target_col_letter: str,
    main_sheet_name: str,
    main_id_col_letter: str,
    main_country_col_letter: str,
    country_lit: str,
) -> str:
    safe_name = main_sheet_name.replace("'", "''")
    main_ref = f"'{safe_name}'!"
    id_range = f"{main_ref}${main_id_col_letter}:${main_id_col_letter}"
    country_range = f"{main_ref}${main_country_col_letter}:${main_country_col_letter}"
    target_range = f"{main_ref}${target_col_letter}:${target_col_letter}"

    match_expr = f"MATCH({id_lit},{id_range},0)"
    country_expr = f"INDEX({country_range},{match_expr})"
    target_expr = f"INDEX({target_range},{match_expr})"

    return (
        f"=IFERROR("
        f"IF({country_expr}={country_lit},"
        f'IF({target_expr}="","",{target_expr}),'
        f'""),'
        f'"")'
    )


_NO_ANSWER_VARIANTS = [
    "No Answer 1",
    "No Answer 2",
    "No Answer 3",
    "No Answer 4",
    "No Answer 5",
    "No Answer 5 up",
]


def _apply_status_color_cf(ws, status_range_str: str) -> None:
    for status_text, (fill_hex, font_hex) in STATUS_COLORS.items():
        rule = CellIsRule(
            operator="equal",
            formula=[f'"{status_text}"'],
            fill=PatternFill(
                "solid",
                start_color=_ff_argb(fill_hex),
                end_color=_ff_argb(fill_hex),
            ),
            font=Font(color=_ff_argb(font_hex), name="Arial"),
        )
        ws.conditional_formatting.add(status_range_str, rule)

    fill_hex, font_hex = NO_ANSWER_COLORS
    for variant in _NO_ANSWER_VARIANTS:
        rule = CellIsRule(
            operator="equal",
            formula=[f'"{variant}"'],
            fill=PatternFill(
                "solid",
                start_color=_ff_argb(fill_hex),
                end_color=_ff_argb(fill_hex),
            ),
            font=Font(color=_ff_argb(font_hex), name="Arial"),
        )
        ws.conditional_formatting.add(status_range_str, rule)


def _apply_cb_black_cf(
    ws,
    cb_range_str: str,
    status_col_letter: str,
    first_data_row: int,
) -> None:
    rule = FormulaRule(
        formula=[
            f'AND(${status_col_letter}{first_data_row}<>"Call Again",'
            f'${status_col_letter}{first_data_row}<>"")'
        ],
        fill=PatternFill("solid", start_color="FF000000", end_color="FF000000"),
    )
    ws.conditional_formatting.add(cb_range_str, rule)


def _apply_comments_yellow_cf(ws, comments_range_str: str) -> None:
    rule = CellIsRule(
        operator="equal",
        formula=['"There was no comments on powerBI"'],
        fill=PatternFill("solid", start_color="FFFFFF00", end_color="FFFFFF00"),
    )
    ws.conditional_formatting.add(comments_range_str, rule)


def _write_pivot_status(
    ws,
    start_row: int,
    start_col: int,
    pivot_name: str,
    rows: list[dict[str, Any]],
    *,
    data_sheet_name: str,
    status_col_letter: str,
    data_first_row: int,
    data_last_row: int,
    main_pivot_sheet: str | None = None,
    main_pivot_status_col_letter: str | None = None,
    main_pivot_status_row_map: dict[str, int] | None = None,
) -> tuple[int, dict[str, int]]:
    seen_lower: set[str] = set()
    ordered_statuses: list[str] = []
    for row in rows:
        status = str(row.get("Status", "") or "").strip()
        status_key = status.lower()
        if status and status_key not in seen_lower:
            seen_lower.add(status_key)
            ordered_statuses.append(status)
    ordered_statuses.sort(key=str.lower)

    label_col = start_col
    status_col = start_col + 1
    count_col = start_col + 2
    pct_col = start_col + 3

    count_letter = get_column_letter(count_col)
    safe_sheet = data_sheet_name.replace("'", "''")
    status_range_ref = (
        f"'{safe_sheet}'!${status_col_letter}${data_first_row}:"
        f"${status_col_letter}${data_last_row}"
    )

    n = len(ordered_statuses)
    total_row = start_row + n
    total_count_ref = f"{count_letter}{total_row}"

    name_cell = ws.cell(row=start_row, column=label_col, value=pivot_name)
    name_cell.font = Font(bold=True, name="Arial", size=11)
    name_cell.alignment = Alignment(horizontal="center", vertical="center")
    name_cell.fill = _make_fill("F4B942")
    ws.merge_cells(
        start_row=start_row,
        start_column=label_col,
        end_row=total_row,
        end_column=label_col,
    )

    status_row_map = {
        status.lower(): start_row + i for i, status in enumerate(ordered_statuses)
    }
    use_main_refs = (
        main_pivot_sheet is not None
        and main_pivot_status_col_letter is not None
        and main_pivot_status_row_map is not None
    )
    safe_main = main_pivot_sheet.replace("'", "''") if main_pivot_sheet else ""

    for i, status in enumerate(ordered_statuses):
        row_i = start_row + i
        if use_main_refs and status.lower() in (main_pivot_status_row_map or {}):
            main_row = main_pivot_status_row_map[status.lower()]  # type: ignore[index]
            main_cell_ref = f"'{safe_main}'!${main_pivot_status_col_letter}${main_row}"
            status_cell = ws.cell(row=row_i, column=status_col, value=f"={main_cell_ref}")
            countif_criterion = main_cell_ref
        else:
            status_cell = ws.cell(row=row_i, column=status_col, value=status)
            countif_criterion = f'"{status.replace(chr(34), chr(34) * 2)}"'

        status_cell.alignment = Alignment(horizontal="center", vertical="center")
        status_cell.font = Font(name="Arial")

        count_cell = ws.cell(
            row=row_i,
            column=count_col,
            value=f"=COUNTIF({status_range_ref},{countif_criterion})",
        )
        count_cell.alignment = Alignment(horizontal="center", vertical="center")
        count_cell.font = Font(name="Arial")

        count_ref = f"{count_letter}{row_i}"
        pct_cell = ws.cell(
            row=row_i,
            column=pct_col,
            value=f"=IFERROR({count_ref}/{total_count_ref},0)",
        )
        pct_cell.number_format = "0%"
        pct_cell.alignment = Alignment(horizontal="center", vertical="center")
        pct_cell.font = Font(name="Arial", bold=True)

    if n > 0:
        pct_letter = get_column_letter(pct_col)
        ws.conditional_formatting.add(
            f"{pct_letter}{start_row}:{pct_letter}{start_row + n - 1}",
            _color_scale_rule(),
        )

    yellow_fill = _make_fill("F4B942")
    gt_status = ws.cell(row=total_row, column=status_col, value="Grand Total")
    if n > 0:
        first_count_ref = f"{count_letter}{start_row}"
        last_count_ref = f"{count_letter}{start_row + n - 1}"
        gt_count = ws.cell(
            row=total_row,
            column=count_col,
            value=f"=SUM({first_count_ref}:{last_count_ref})",
        )
    else:
        gt_count = ws.cell(row=total_row, column=count_col, value=0)

    ws.merge_cells(
        start_row=total_row,
        start_column=count_col,
        end_row=total_row,
        end_column=pct_col,
    )
    gt_status.fill = yellow_fill
    gt_count.fill = yellow_fill
    ws.cell(row=total_row, column=pct_col).fill = yellow_fill

    for cell in (gt_status, gt_count):
        cell.font = Font(bold=True, name="Arial")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_num in range(start_row, total_row + 1):
        for col_num in range(label_col, pct_col + 1):
            ws.cell(row=row_num, column=col_num).border = THIN_BORDER

    return total_row + 1, status_row_map


def _write_pivot_call_attempts(
    ws,
    start_row: int,
    start_col: int,
    rows: list[dict[str, Any]],
    *,
    data_sheet_name: str,
    call_attempts_col_letter: str,
    data_first_row: int,
    data_last_row: int,
) -> int:
    del rows
    bucket_order = ["1", "2", "3", "4", "5+"]
    label_col = start_col
    bucket_col = start_col + 1
    count_col = start_col + 2
    pct_col = start_col + 3

    count_letter = get_column_letter(count_col)
    safe_sheet = data_sheet_name.replace("'", "''")
    ca_range_ref = (
        f"'{safe_sheet}'!${call_attempts_col_letter}${data_first_row}:"
        f"${call_attempts_col_letter}${data_last_row}"
    )

    n = len(bucket_order)
    total_row = start_row + n
    total_count_ref = f"{count_letter}{total_row}"

    name_cell = ws.cell(row=start_row, column=label_col, value="Call Attempts")
    name_cell.font = Font(bold=True, name="Arial", size=11)
    name_cell.alignment = Alignment(horizontal="center", vertical="center")
    name_cell.fill = _make_fill("F4B942")
    ws.merge_cells(
        start_row=start_row,
        start_column=label_col,
        end_row=total_row,
        end_column=label_col,
    )

    for i, bucket in enumerate(bucket_order):
        row_i = start_row + i
        bucket_cell = ws.cell(row=row_i, column=bucket_col, value=bucket)
        bucket_cell.alignment = Alignment(horizontal="center", vertical="center")
        bucket_cell.font = Font(name="Arial")

        criterion = '">=5"' if bucket == "5+" else bucket
        count_cell = ws.cell(
            row=row_i,
            column=count_col,
            value=f"=COUNTIF({ca_range_ref},{criterion})",
        )
        count_cell.alignment = Alignment(horizontal="center", vertical="center")
        count_cell.font = Font(name="Arial")

        count_ref = f"{count_letter}{row_i}"
        pct_cell = ws.cell(
            row=row_i,
            column=pct_col,
            value=f"=IFERROR({count_ref}/{total_count_ref},0)",
        )
        pct_cell.number_format = "0%"
        pct_cell.alignment = Alignment(horizontal="center", vertical="center")
        pct_cell.font = Font(name="Arial", bold=True)

    pct_letter = get_column_letter(pct_col)
    ws.conditional_formatting.add(
        f"{pct_letter}{start_row}:{pct_letter}{start_row + n - 1}",
        _color_scale_rule(),
    )

    yellow_fill = _make_fill("F4B942")
    first_count_ref = f"{count_letter}{start_row}"
    last_count_ref = f"{count_letter}{start_row + n - 1}"
    gt_bucket = ws.cell(row=total_row, column=bucket_col, value="Grand total")
    gt_count = ws.cell(
        row=total_row,
        column=count_col,
        value=f"=SUM({first_count_ref}:{last_count_ref})",
    )

    ws.merge_cells(
        start_row=total_row,
        start_column=count_col,
        end_row=total_row,
        end_column=pct_col,
    )
    gt_bucket.fill = yellow_fill
    gt_count.fill = yellow_fill
    ws.cell(row=total_row, column=pct_col).fill = yellow_fill

    for cell in (gt_bucket, gt_count):
        cell.font = Font(bold=True, name="Arial")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_num in range(start_row, total_row + 1):
        for col_num in range(label_col, pct_col + 1):
            ws.cell(row=row_num, column=col_num).border = THIN_BORDER

    return total_row + 1


def _write_data_table_and_pivots(
    ws,
    rows: list[dict[str, Any]],
    pivot_label: str,
    table_name: str,
    include_call_attempts: bool,
    link_to_main: str | None = None,
    country_filter: str | None = None,
    main_pivot_info: dict[str, Any] | None = None,
) -> dict[str, Any] | None:
    status_col_idx = OUTPUT_COLUMNS.index("Status") + 1
    cb_col_idx = OUTPUT_COLUMNS.index("CB") + 1
    comments_col_idx = OUTPUT_COLUMNS.index("Comments") + 1
    id_col_idx = OUTPUT_COLUMNS.index("ID") + 1
    country_col_idx = OUTPUT_COLUMNS.index("Country") + 1
    created_col_idx = OUTPUT_COLUMNS.index("Created") + 1

    formula_mode = link_to_main is not None and country_filter is not None
    if formula_mode:
        country_lit = _str_literal_for_formula(country_filter)
        main_id_letter = get_column_letter(id_col_idx)
        main_country_letter = get_column_letter(country_col_idx)
        created_number_format = "yyyy-mm-dd hh:mm:ss"

    ws.append(OUTPUT_COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True, name="Arial")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=1, column=comments_col_idx).alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )

    if formula_mode:
        for idx, row in enumerate(rows):
            excel_row = idx + 2
            id_lit = _id_literal_for_formula(row.get("ID"))
            status_norm = normalize_status(row.get("Status", ""))

            for col_idx in range(1, len(OUTPUT_COLUMNS) + 1):
                formula = _country_cell_formula(
                    id_lit=id_lit,
                    target_col_letter=get_column_letter(col_idx),
                    main_sheet_name=link_to_main,
                    main_id_col_letter=main_id_letter,
                    main_country_col_letter=main_country_letter,
                    country_lit=country_lit,
                )
                cell = ws.cell(row=excel_row, column=col_idx, value=formula)
                if col_idx == comments_col_idx:
                    cell.alignment = Alignment(
                        wrap_text=True,
                        vertical="center",
                        horizontal="center",
                    )
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(name="Arial")
                if col_idx == created_col_idx:
                    cell.number_format = created_number_format

            status_cell = ws.cell(row=excel_row, column=status_col_idx)
            if NO_ANSWER_STATUS_RE.match(status_norm):
                fill_hex, font_hex = NO_ANSWER_COLORS
            else:
                fill_hex, font_hex = STATUS_COLORS.get(status_norm, (None, None))
            if fill_hex:
                status_cell.fill = _make_fill(fill_hex)
                status_cell.font = _make_font(font_hex)

            if row.get("CB") is None:
                ws.cell(row=excel_row, column=cb_col_idx).fill = _make_fill("000000")
            if row.get("_comments_yellow"):
                ws.cell(row=excel_row, column=comments_col_idx).fill = _make_fill("FFFF00")
    else:
        for row in rows:
            data = [row.get(column, "") for column in OUTPUT_COLUMNS]
            cb_index = OUTPUT_COLUMNS.index("CB")
            if data[cb_index] is None:
                data[cb_index] = ""

            ws.append(data)
            excel_row = ws.max_row
            status_norm = normalize_status(row.get("Status", ""))

            status_cell = ws.cell(row=excel_row, column=status_col_idx)
            if NO_ANSWER_STATUS_RE.match(status_norm):
                fill_hex, font_hex = NO_ANSWER_COLORS
            else:
                fill_hex, font_hex = STATUS_COLORS.get(status_norm, (None, None))
            if fill_hex:
                status_cell.fill = _make_fill(fill_hex)
                status_cell.font = _make_font(font_hex)

            if row.get("CB") is None:
                ws.cell(row=excel_row, column=cb_col_idx).fill = _make_fill("000000")

            comments_cell = ws.cell(row=excel_row, column=comments_col_idx)
            comments_cell.alignment = Alignment(
                wrap_text=True,
                vertical="center",
                horizontal="center",
            )
            if row.get("_comments_yellow"):
                comments_cell.fill = _make_fill("FFFF00")

            for col_idx in range(1, len(OUTPUT_COLUMNS) + 1):
                cell = ws.cell(row=excel_row, column=col_idx)
                if col_idx != comments_col_idx:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                rgb = (
                    cell.font.color.rgb
                    if cell.font.color and cell.font.color.type == "rgb"
                    else "FF000000"
                )
                cell.font = Font(name="Arial", color=rgb, bold=cell.font.bold)

    last_data_row = ws.max_row
    _apply_all_borders(ws, last_data_row, len(OUTPUT_COLUMNS))

    if formula_mode and last_data_row >= 2:
        status_letter = get_column_letter(status_col_idx)
        cb_letter = get_column_letter(cb_col_idx)
        comments_letter = get_column_letter(comments_col_idx)
        _apply_status_color_cf(ws, f"{status_letter}2:{status_letter}{last_data_row}")
        _apply_cb_black_cf(
            ws,
            f"{cb_letter}2:{cb_letter}{last_data_row}",
            status_letter,
            first_data_row=2,
        )
        _apply_comments_yellow_cf(
            ws,
            f"{comments_letter}2:{comments_letter}{last_data_row}",
        )

    last_col_letter = get_column_letter(len(OUTPUT_COLUMNS))
    table = Table(displayName=table_name, ref=f"A1:{last_col_letter}{last_data_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    for col_idx in range(1, len(OUTPUT_COLUMNS) + 1):
        col_letter = get_column_letter(col_idx)
        column_name = OUTPUT_COLUMNS[col_idx - 1]
        max_length = len(str(column_name))

        for row in rows[:100]:
            value = row.get(column_name)
            if value is None or value == "":
                continue
            if isinstance(value, _dt.datetime):
                length = 19
            elif isinstance(value, _dt.date):
                length = 10
            else:
                length = len(str(value))
            max_length = max(max_length, length)

        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

    pivot_row = last_data_row + PIVOT_GAP_ROWS
    pivot_col = PIVOT_START_COL
    status_letter = get_column_letter(status_col_idx)
    call_att_letter = get_column_letter(OUTPUT_COLUMNS.index("Call Attempts") + 1)
    pivot_status_col_letter = get_column_letter(pivot_col + 1)

    main_sheet = main_pivot_info["sheet_name"] if main_pivot_info else None
    main_col_letter = main_pivot_info["status_col_letter"] if main_pivot_info else None
    main_row_map = main_pivot_info["status_row_map"] if main_pivot_info else None

    next_row, status_row_map = _write_pivot_status(
        ws,
        pivot_row,
        pivot_col,
        pivot_label,
        rows,
        data_sheet_name=ws.title,
        status_col_letter=status_letter,
        data_first_row=2,
        data_last_row=last_data_row,
        main_pivot_sheet=main_sheet,
        main_pivot_status_col_letter=main_col_letter,
        main_pivot_status_row_map=main_row_map,
    )

    if include_call_attempts:
        next_row += PIVOT_INTER_GAP
        _write_pivot_call_attempts(
            ws,
            next_row,
            pivot_col,
            rows,
            data_sheet_name=ws.title,
            call_attempts_col_letter=call_att_letter,
            data_first_row=2,
            data_last_row=last_data_row,
        )

    if not formula_mode:
        return {
            "sheet_name": ws.title,
            "status_col_letter": pivot_status_col_letter,
            "status_row_map": status_row_map,
        }
    return None


_INVALID_SHEET_CHARS_RE = re.compile(r"[\\/*?\[\]:]")


def _sanitize_sheet_name(name: str, used: set[str]) -> str:
    sanitized = _INVALID_SHEET_CHARS_RE.sub("_", str(name).strip())
    sanitized = sanitized.strip("'")
    sanitized = sanitized[:31] or "Sheet"

    base = sanitized
    counter = 1
    used_lower = {item.lower() for item in used}
    while sanitized.lower() in used_lower:
        suffix = f"_{counter}"
        sanitized = (base[: 31 - len(suffix)]) + suffix
        counter += 1

    used.add(sanitized)
    return sanitized


def _sanitize_table_name(name: str, used: set[str]) -> str:
    sanitized = re.sub(r"[^A-Za-z0-9_]", "_", str(name))
    if not sanitized or sanitized[0].isdigit():
        sanitized = f"T_{sanitized}"
    if sanitized == "_":
        sanitized = "T_Table"

    base = sanitized
    counter = 1
    while sanitized in used:
        sanitized = f"{base}_{counter}"
        counter += 1

    used.add(sanitized)
    return sanitized


def write_output(rows: list[dict[str, Any]], output_file: Path) -> None:
    workbook = Workbook()
    used_sheet_names: set[str] = set()
    used_table_names: set[str] = set()

    rows = sorted(rows, key=lambda row: str(row.get("Status", "") or "").strip().lower())

    ws_main = workbook.active
    ws_main.title = _sanitize_sheet_name("Main Report", used_sheet_names)
    main_pivot_info = _write_data_table_and_pivots(
        ws_main,
        rows,
        pivot_label=MAIN_REPORT_PIVOT_LABEL,
        table_name=_sanitize_table_name("CRMOutput_Main", used_table_names),
        include_call_attempts=True,
    )

    countries: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        country = str(row.get("Country", "") or "").strip()
        if country:
            countries[country].append(row)

    for country in sorted(countries.keys(), key=str.lower):
        sheet_name = _sanitize_sheet_name(country, used_sheet_names)
        ws_country = workbook.create_sheet(title=sheet_name)
        _write_data_table_and_pivots(
            ws_country,
            countries[country],
            pivot_label=country,
            table_name=_sanitize_table_name(f"CRMOutput_{country}", used_table_names),
            include_call_attempts=False,
            link_to_main=ws_main.title,
            country_filter=country,
            main_pivot_info=main_pivot_info,
        )

    output_file.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_file)


def build_output(
    powerbi_report: Path,
    crm_files: list[Path],
    platforms: list[str],
    output_file: Path,
    powerbi_sheet: str | None = None,
    crm_sheet: str | None = None,
) -> None:
    if len(crm_files) != len(platforms):
        raise ValueError("Each CRM file must have exactly one platform name.")

    powerbi_lookup = read_powerbi_lookup(powerbi_report, powerbi_sheet)
    all_rows: list[dict[str, Any]] = []

    for crm_file, platform in zip(crm_files, platforms):
        all_rows.extend(read_crm_rows(crm_file, platform, powerbi_lookup, crm_sheet))

    write_output(all_rows, output_file)


def prompt(message: str, allow_empty: bool = False) -> str:
    while True:
        value = input(message).strip()
        if value or allow_empty:
            return value
        print("  This field cannot be empty. Please try again.", file=sys.stderr)


def prompt_path(message: str, must_exist: bool = True) -> Path:
    while True:
        raw = prompt(message)
        path = Path(raw)
        if must_exist and not path.exists():
            print(
                f"  File not found: {path}. Please check the path and try again.",
                file=sys.stderr,
            )
            continue
        return path


def prompt_optional(message: str) -> str | None:
    value = input(message).strip()
    return value if value else None


def collect_inputs() -> dict[str, Any]:
    print("\n=== CRM + PowerBI Country Report Tool ===\n")

    powerbi_report = prompt_path("PowerBI report file path: ")
    powerbi_sheet = prompt_optional(
        "  Sheet name (leave blank to use the active sheet): "
    )

    crm_files: list[Path] = []
    platforms: list[str] = []

    print("\nEnter CRM files one at a time. Leave the path blank when done.")
    while True:
        index = len(crm_files) + 1
        crm_path_input = input(
            f"CRM file #{index} path (or press Enter to finish): "
        ).strip()

        if not crm_path_input:
            if not crm_files:
                print("  You must provide at least one CRM file.", file=sys.stderr)
                continue
            break

        crm_path = Path(crm_path_input)
        if not crm_path.exists():
            print(
                f"  File not found: {crm_path}. Please check the path and try again.",
                file=sys.stderr,
            )
            continue

        platform = prompt(f"  Platform name for '{crm_path.name}': ")
        crm_files.append(crm_path)
        platforms.append(platform)

    crm_sheet = prompt_optional(
        "\nCRM sheet name for all CRM files (leave blank for each file's active sheet): "
    )

    output_raw = input(
        f"\nOutput file path (leave blank for '{DEFAULT_OUTPUT_FILENAME}'): "
    ).strip()
    output_file = Path(output_raw) if output_raw else Path(DEFAULT_OUTPUT_FILENAME)

    return {
        "powerbi_report": powerbi_report,
        "powerbi_sheet": powerbi_sheet,
        "crm_files": crm_files,
        "platforms": platforms,
        "crm_sheet": crm_sheet,
        "output_file": output_file,
    }


def main() -> int:
    try:
        inputs = collect_inputs()
        print("\nProcessing...")
        build_output(
            powerbi_report=inputs["powerbi_report"],
            crm_files=inputs["crm_files"],
            platforms=inputs["platforms"],
            output_file=inputs["output_file"],
            powerbi_sheet=inputs["powerbi_sheet"],
            crm_sheet=inputs["crm_sheet"],
        )
    except KeyboardInterrupt:
        print("\nCancelled.", file=sys.stderr)
        return 1
    except Exception as exc:
        print(f"\nError: {exc}", file=sys.stderr)
        return 1

    print(f"\nDone! Output written to: {inputs['output_file']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
