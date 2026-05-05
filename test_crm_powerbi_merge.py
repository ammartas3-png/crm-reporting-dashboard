import unittest
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import Workbook, load_workbook

from crm_powerbi_merge import build_output, comments_for_status, extract_comments


def save_workbook(path: Path, headers: list[str], rows: list[list[object]]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    workbook.save(path)


def read_output_rows(path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(path)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    headers = rows[0]
    return [dict(zip(headers, row)) for row in rows[1:]]


class CrmPowerbiMergeTest(unittest.TestCase):
    def test_extract_comments_returns_bottom_to_top_order(self) -> None:
        comments = (
            "2026-05-05 13:47 | Ismail Al | na vm;\n"
            "2026-05-05 13:15 | Ismail Al | na;\n"
            "2026-05-05 12:52 | Ismail Al | 2x na vm;\n"
            "2026-05-05 12:50 | Ismail Al | In progress;"
        )

        self.assertEqual(
            extract_comments(comments),
            "In progress\n2x na vm\nna\nna vm",
        )

    def test_comments_for_status_business_rules(self) -> None:
        self.assertEqual(comments_for_status("DNC", "real comment"), "")
        self.assertEqual(comments_for_status("Invalid country", "real comment"), "")
        self.assertEqual(comments_for_status("No Answer 3", "real comment"), "NA VM")
        self.assertEqual(comments_for_status("New", "real comment"), "real comment")

    def test_build_output_matches_crm_to_powerbi(self) -> None:
        with TemporaryDirectory() as temp_dir:
            tmp_path = Path(temp_dir)
            powerbi = tmp_path / "powerbi.xlsx"
            crm_a = tmp_path / "crm_a.xlsx"
            crm_b = tmp_path / "crm_b.xlsx"
            output = tmp_path / "output.xlsx"

            save_workbook(
                powerbi,
                [
                    "Account No",
                    "Brand name",
                    "Last 10 Comments",
                    "Voip Calls Attempts Cnt",
                ],
                [
                    [
                        "1001",
                        "Brand A",
                        "2026-05-05 | Agent | second;\n"
                        "2026-05-04 | Agent | first;",
                        4,
                    ],
                    ["1002", "Brand A", "2026-05-05 | Agent | should hide;", 2],
                    [
                        "1003",
                        "Brand B",
                        "2026-05-05 | Agent | should become NA VM;",
                        1,
                    ],
                    ["1001", "Brand B", "2026-05-05 | Agent | wrong brand;", 9],
                ],
            )
            save_workbook(
                crm_a,
                [
                    "Customer Type",
                    "ID",
                    "Created",
                    "Name",
                    "Department",
                    "Status",
                    "Country",
                    "Assigned to",
                ],
                [
                    ["Lead", "1001", "2026-05-01", "Alice", "Sales", "New", "DE", "Sam"],
                    ["Lead", "1002", "2026-05-02", "Bob", "Sales", "DNC", "DE", "Sam"],
                ],
            )
            save_workbook(
                crm_b,
                [
                    "Customer Type",
                    "ID",
                    "Created",
                    "Name",
                    "Department",
                    "Status",
                    "Country",
                    "Assigned to",
                ],
                [
                    [
                        "Lead",
                        "1003",
                        "2026-05-03",
                        "Cara",
                        "Sales",
                        "No Answer 5",
                        "FR",
                        "Jo",
                    ],
                ],
            )

            build_output(powerbi, [crm_a, crm_b], ["Brand A", "Brand B"], output)

            rows = read_output_rows(output)
            self.assertEqual(
                list(rows[0]),
                [
                    "Platform",
                    "Customer Type",
                    "ID",
                    "Created",
                    "Name",
                    "Department",
                    "Status",
                    "Country",
                    "Assigned to",
                    "Comments",
                    "Call Attempts",
                ],
            )
            self.assertEqual(rows[0]["Platform"], "Brand A")
            self.assertEqual(rows[0]["Comments"], "first\nsecond")
            self.assertEqual(rows[0]["Call Attempts"], 4)
            self.assertIsNone(rows[1]["Comments"])
            self.assertEqual(rows[1]["Call Attempts"], 2)
            self.assertEqual(rows[2]["Platform"], "Brand B")
            self.assertEqual(rows[2]["Comments"], "NA VM")
            self.assertEqual(rows[2]["Call Attempts"], 1)


if __name__ == "__main__":
    unittest.main()
