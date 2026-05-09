"""Create a CRM output workbook enriched with PowerBI comments and call attempts.

The module contains the report-generation logic used by both the web UI and an
optional command-line entry point.
"""

from __future__ import annotations

import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


CRM_COLUMNS = [
    "Customer Type",
    "ID",
    "Created",
    "Name",
    "Department",
    "Status",
    "Country",
    "Campaign",
    "Sub-Campaign",
    "Placement",
    "Assigned to",
]

POWERBI_COLUMNS = [
    "Account No",
    "Brand",
    "Last 10 Comments",
    "Voip Calls Attempts Cnt",
]

OUTPUT_COLUMNS = [
    "Platform",
    "Customer Type",
    "ID",
    "Created",
    "Name",
    "Department",
    "Status",
    "CB",
    "Country",
    "Campaign",
    "Sub-Campaign",
    "Placement",
    "Assigned to",
    "Comments",
    "Call Attempts",
]

STATUS_COLORS: dict[str, tuple[str, str]] = {
    "call again": ("FFFF00", "000000"),
    "decline": ("FF0000", "FFFFFF"),
    "denied registration": ("ADD8E6", "000000"),
    "duplicate": ("4B0082", "FFFFFF"),
    "no interest": ("8B0000", "FFFFFF"),
    "no language": ("D3D3D3", "000000"),
    "no potential": ("FFB6C1", "000000"),
    "no potential - no bank account": ("0000FF", "000000"),
    "no potential - no documents": ("FFD580", "000000"),
    "not interested": ("8B0000", "FFFFFF"),
    "potential": ("90EE90", "000000"),
    "recall": ("404040", "FFFFFF"),
    "telemarketing": ("006400", "FFFFFF"),
    "under 18": ("FFA500", "000000"),
    "wrong number or email": ("D3D3D3", "000000"),
}

NO_ANSWER_STATUS_RE = re.compile(r"no answer ([1-5]|5\s*up)", re.IGNORECASE)
NO_ANSWER_COLORS = ("FF9999", "000000")

NO_COMMENT_STATUSES = {
    "dnc",
    "invalid country",
    "wrong number or email",
    "duplicate",
    "no potential - no documents",
    "test",
    "under 18",
    "no language",
}

STATUS_LIST: list[str] = sorted(
    [
        "Call Again",
        "Decline",
        "Denied Registration",
        "Duplicate",
        "No Answer 1-5",
        "No Answer 5 up",
        "No Interest",
        "No Language",
        "No Potential",
        "No Potential - No Bank Account",
        "No Potential - No Documents",
        "Not Interested",
        "Potential",
        "Recall",
        "Telemarketing",
        "Under 18",
        "Wrong Number or Email",
    ],
    key=str.lower,
)

COMMENT_RE = re.compile(r"\|\s*([^|;]*?)\s*;")
NA_LIKE_PATTERN = re.compile(r"^(na(\s+vm)?|vm|dvm)$", re.IGNORECASE)
STATUS_COMMENT_RE = re.compile(
    r"^("
    r"call again|decline|denied registration|duplicate"
    r"|no answer [1-5]|no answer 5\s*up"
    r"|no language|no potential|no potential\s*[-–]\s*no bank account"
    r"|no potential\s*[-–]\s*no documents"
    r"|not interested|no interest"
    r"|in progress"
    r"|potential|recall|under 18|wrong number or email"
    r")$",
    re.IGNORECASE,
)
LV_PREFIX_RE = re.compile(r"^[lv][1-5]\s+", re.IGNORECASE)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def normalize_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


def normalize_status(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


def normalize_match_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"\s+", " ", str(value).strip()).casefold()


def clean_comment(value: str) -> str:
    return " ".join(value.split())


def extract_comments(last_10_comments: Any) -> list[str]:
    if last_10_comments is None:
        return []
    comments = [
        clean_comment(match.group(1))
        for match in COMMENT_RE.finditer(str(last_10_comments))
    ]
    comments = [comment for comment in comments if comment]
    comments.reverse()
    return comments


def _is_na_like(text: str) -> bool:
    return bool(NA_LIKE_PATTERN.match(text.strip()))


def _preprocess_comment(text: str) -> str | None:
    if text.lower().startswith("email"):
        return None
    if STATUS_COMMENT_RE.match(text.strip()):
        return None
    text = LV_PREFIX_RE.sub("", text).strip()
    return text if text else None


def format_comments(comments: list[str]) -> str:
    preprocessed: list[str] = []
    for comment in comments:
        result = _preprocess_comment(comment)
        if result is not None:
            preprocessed.append(result)

    merged: list[str] = []
    i = 0
    while i < len(preprocessed):
        if _is_na_like(preprocessed[i]):
            run = 1
            while i + run < len(preprocessed) and _is_na_like(preprocessed[i + run]):
                run += 1
            merged.append(f"NA VM x{run}" if run > 1 else "NA VM")
            i += run
        else:
            merged.append(preprocessed[i])
            i += 1

    return " // ".join(merged)


def comments_for_status(
    status: Any,
    matched_comments: list[str],
    found_in_powerbi: bool,
) -> tuple[str, bool]:
    normalized = normalize_status(status)

    if NO_ANSWER_STATUS_RE.match(normalized):
        return "NA VM", False

    if normalized in NO_COMMENT_STATUSES:
        return "", False

    if not found_in_powerbi:
        return "There was no comments on powerBI", True

    text = format_comments(matched_comments)
    if not text:
        return "There was no comments on powerBI", True

    return text, False


def worksheet_from_file(path: Path, sheet_name: str | None = None):
    workbook = load_workbook(path, data_only=True)
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"Sheet {sheet_name!r} not found in {path.name}. "
                f"Available sheets: {', '.join(workbook.sheetnames)}"
            )
        return workbook[sheet_name]
    return workbook.active


def header_indexes(
    headers: Iterable[Any],
    required_columns: list[str],
    path: Path,
) -> dict[str, int]:
    normalized_headers = {normalize_header(header): i for i, header in enumerate(headers)}
    missing = [
        column
        for column in required_columns
        if normalize_header(column) not in normalized_headers
    ]
    if missing:
        raise ValueError(f"{path.name} is missing required columns: {', '.join(missing)}")
    return {
        column: normalized_headers[normalize_header(column)]
        for column in required_columns
    }


def read_powerbi_lookup(
    powerbi_report: Path,
    sheet_name: str | None = None,
) -> dict[tuple[str, str], dict[str, Any]]:
    worksheet = worksheet_from_file(powerbi_report, sheet_name)
    rows = worksheet.iter_rows(values_only=True)

    try:
        headers = next(rows)
    except StopIteration as exc:
        raise ValueError(f"{powerbi_report.name} is empty") from exc

    indexes = header_indexes(headers, POWERBI_COLUMNS, powerbi_report)
    lookup: dict[tuple[str, str], dict[str, Any]] = {}

    for row in rows:
        account_no = row[indexes["Account No"]]
        brand_name = row[indexes["Brand"]]
        account_key = normalize_match_value(account_no)
        brand_key = normalize_match_value(brand_name)

        if not account_key or not brand_key:
            continue

        lookup[(account_key, brand_key)] = {
            "Comments": extract_comments(row[indexes["Last 10 Comments"]]),
            "Call Attempts": row[indexes["Voip Calls Attempts Cnt"]],
        }

    return lookup


def read_crm_rows(
    crm_file: Path,
    platform: str,
    powerbi_lookup: dict[tuple[str, str], dict[str, Any]],
    sheet_name: str | None = None,
) -> list[dict[str, Any]]:
    worksheet = worksheet_from_file(crm_file, sheet_name)
    rows = worksheet.iter_rows(values_only=True)

    try:
        headers = next(rows)
    except StopIteration as exc:
        raise ValueError(f"{crm_file.name} is empty") from exc

    indexes = header_indexes(headers, CRM_COLUMNS, crm_file)
    platform_key = normalize_match_value(platform)
    output_rows: list[dict[str, Any]] = []

    for row in rows:
        if all(value is None for value in row):
            continue

        crm_values = {column: row[indexes[column]] for column in CRM_COLUMNS}

        customer_type_norm = normalize_status(crm_values.get("Customer Type", ""))
        if customer_type_norm == "depositor":
            crm_values["Status"] = "Telemarketing"
            output_rows.append(
                {
                    "Platform": platform,
                    **crm_values,
                    "CB": None,
                    "Comments": "",
                    "_comments_yellow": False,
                    "Call Attempts": 1,
                }
            )
            continue

        account_key = normalize_match_value(crm_values["ID"])
        matched = powerbi_lookup.get((account_key, platform_key))
        found_in_powerbi = matched is not None
        matched_comments: list[str] = matched.get("Comments", []) if matched else []

        comment_text, use_yellow = comments_for_status(
            crm_values["Status"],
            matched_comments,
            found_in_powerbi,
        )

        status_norm = normalize_status(crm_values["Status"])
        cb_value = "" if status_norm == "call again" else None

        call_attempts = matched.get("Call Attempts") if matched else None
        if call_attempts in (None, ""):
            call_attempts = 1

        output_rows.append(
            {
                "Platform": platform,
                **crm_values,
                "CB": cb_value,
                "Comments": comment_text,
                "_comments_yellow": use_yellow,
                "Call Attempts": call_attempts,
            }
        )

    return output_rows


PCT_COLOR_SCALE = ColorScaleRule(
    start_type="num",
    start_value=0,
    start_color="F8696B",
    mid_type="num",
    mid_value=0.5,
    mid_color="FFEB84",
    end_type="num",
    end_value=1,
    end_color="63BE7B",
)


def _write_pivot_status(
    ws,
    start_row: int,
    start_col: int,
    pivot_name: str,
    rows: list[dict[str, Any]],
    data_sheet_name: str = "CRM Output",
    filter_country: str | None = None,
    main_sheet_name: str | None = None,
) -> int:
    del rows
    label_col = start_col
    status_col = start_col + 1
    count_col = start_col + 2
    pct_col = start_col + 3

    ref_sheet = main_sheet_name if (filter_country and main_sheet_name) else data_sheet_name
    status_letter = get_column_letter(OUTPUT_COLUMNS.index("Status") + 1)
    status_range = f"'{ref_sheet}'!{status_letter}:{status_letter}"

    if filter_country and main_sheet_name:
        country_letter = get_column_letter(OUTPUT_COLUMNS.index("Country") + 1)
        country_range = f"'{main_sheet_name}'!{country_letter}:{country_letter}"
        safe_country = filter_country.replace('"', '""')

    n = len(STATUS_LIST)
    count_col_letter = get_column_letter(count_col)
    first_count_row = start_row
    last_count_row = start_row + n - 1
    total_row = start_row + n
    total_cell_addr = f"{count_col_letter}{total_row}"

    name_cell = ws.cell(row=start_row, column=label_col, value=pivot_name)
    name_cell.font = Font(bold=True, name="Arial", size=11)
    name_cell.alignment = Alignment(horizontal="center", vertical="center")
    name_cell.fill = PatternFill("solid", start_color="F4B942", fgColor="F4B942")
    if n > 1:
        ws.merge_cells(
            start_row=start_row,
            start_column=label_col,
            end_row=total_row,
            end_column=label_col,
        )

    for i, status in enumerate(STATUS_LIST):
        row_i = start_row + i

        status_cell = ws.cell(row=row_i, column=status_col, value=status)
        status_cell.alignment = Alignment(horizontal="center", vertical="center")
        status_cell.font = Font(name="Arial")

        safe_status = status.replace('"', '""')
        if filter_country and main_sheet_name:
            count_formula = (
                f'=COUNTIFS({status_range},"{safe_status}",'
                f'{country_range},"{safe_country}")'
            )
        else:
            count_formula = f'=COUNTIF({status_range},"{safe_status}")'

        count_cell = ws.cell(row=row_i, column=count_col, value=count_formula)
        count_cell.alignment = Alignment(horizontal="center", vertical="center")
        count_cell.font = Font(name="Arial")

        count_cell_addr = f"{count_col_letter}{row_i}"
        pct_cell = ws.cell(
            row=row_i,
            column=pct_col,
            value=f"=IFERROR({count_cell_addr}/{total_cell_addr},0)",
        )
        pct_cell.number_format = "0%"
        pct_cell.alignment = Alignment(horizontal="center", vertical="center")
        pct_cell.font = Font(name="Arial", bold=True)

    gt_label = ws.cell(row=total_row, column=label_col)
    gt_status = ws.cell(row=total_row, column=status_col, value="Total")
    gt_count = ws.cell(
        row=total_row,
        column=count_col,
        value=f"=SUM({count_col_letter}{first_count_row}:{count_col_letter}{last_count_row})",
    )
    gt_pct = ws.cell(row=total_row, column=pct_col, value="")

    gt_label.font = Font(bold=True, name="Arial")
    gt_label.alignment = Alignment(horizontal="center", vertical="center")
    gt_status.font = Font(bold=True, name="Arial")
    gt_status.alignment = Alignment(horizontal="center", vertical="center")
    gt_status.fill = PatternFill("solid", start_color="F4B942", fgColor="F4B942")

    ws.merge_cells(
        start_row=total_row,
        start_column=count_col,
        end_row=total_row,
        end_column=pct_col,
    )
    for cell in (gt_count, gt_pct):
        cell.font = Font(bold=True, name="Arial")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", start_color="F4B942", fgColor="F4B942")

    for row_num in range(start_row, total_row + 1):
        for col_num in range(label_col, pct_col + 1):
            ws.cell(row=row_num, column=col_num).border = THIN_BORDER

    pct_col_letter = get_column_letter(pct_col)
    ws.conditional_formatting.add(
        f"{pct_col_letter}{start_row}:{pct_col_letter}{last_count_row}",
        PCT_COLOR_SCALE,
    )

    return total_row + 1


def _write_pivot_call_attempts(
    ws,
    start_row: int,
    start_col: int,
    rows: list[dict[str, Any]],
    data_sheet_name: str = "CRM Output",
    filter_country: str | None = None,
    main_sheet_name: str | None = None,
) -> int:
    del rows
    bucket_order = ["1", "2", "3", "4", "5+"]

    label_col = start_col
    bucket_col = start_col + 1
    count_col = start_col + 2
    pct_col = start_col + 3

    ref_sheet = main_sheet_name if (filter_country and main_sheet_name) else data_sheet_name
    ca_letter = get_column_letter(OUTPUT_COLUMNS.index("Call Attempts") + 1)
    ca_range = f"'{ref_sheet}'!{ca_letter}:{ca_letter}"

    if filter_country and main_sheet_name:
        country_letter = get_column_letter(OUTPUT_COLUMNS.index("Country") + 1)
        country_range = f"'{main_sheet_name}'!{country_letter}:{country_letter}"
        safe_country = filter_country.replace('"', '""')

    count_col_letter = get_column_letter(count_col)
    n = len(bucket_order)
    first_count_row = start_row
    last_count_row = start_row + n - 1
    total_row = start_row + n
    total_cell_addr = f"{count_col_letter}{total_row}"

    name_cell = ws.cell(row=start_row, column=label_col, value="Call Attempts")
    name_cell.font = Font(bold=True, name="Arial", size=11)
    name_cell.alignment = Alignment(horizontal="center", vertical="center")
    name_cell.fill = PatternFill("solid", start_color="F4B942", fgColor="F4B942")
    if n > 1:
        ws.merge_cells(
            start_row=start_row,
            start_column=label_col,
            end_row=total_row,
            end_column=label_col,
        )

    for i, bucket in enumerate(bucket_order):
        row_i = start_row + i

        bucket_cell = ws.cell(row=row_i, column=bucket_col, value=bucket)
        bucket_cell.alignment = Alignment(horizontal="center", vertical="center")
        bucket_cell.font = Font(name="Arial")

        criterion = '">=5"' if bucket == "5+" else bucket
        if filter_country and main_sheet_name:
            count_formula = (
                f"=COUNTIFS({ca_range},{criterion},"
                f'{country_range},"{safe_country}")'
            )
        else:
            count_formula = f"=COUNTIF({ca_range},{criterion})"

        count_cell = ws.cell(row=row_i, column=count_col, value=count_formula)
        count_cell.alignment = Alignment(horizontal="center", vertical="center")
        count_cell.font = Font(name="Arial")

        count_cell_addr = f"{count_col_letter}{row_i}"
        pct_cell = ws.cell(
            row=row_i,
            column=pct_col,
            value=f"=IFERROR({count_cell_addr}/{total_cell_addr},0)",
        )
        pct_cell.number_format = "0%"
        pct_cell.alignment = Alignment(horizontal="center", vertical="center")
        pct_cell.font = Font(name="Arial", bold=True)

    gt_label = ws.cell(row=total_row, column=label_col)
    gt_bucket = ws.cell(row=total_row, column=bucket_col, value="Total")
    gt_count = ws.cell(
        row=total_row,
        column=count_col,
        value=f"=SUM({count_col_letter}{first_count_row}:{count_col_letter}{last_count_row})",
    )
    gt_pct = ws.cell(row=total_row, column=pct_col, value="")

    gt_label.font = Font(bold=True, name="Arial")
    gt_label.alignment = Alignment(horizontal="center", vertical="center")
    gt_bucket.font = Font(bold=True, name="Arial")
    gt_bucket.alignment = Alignment(horizontal="center", vertical="center")
    gt_bucket.fill = PatternFill("solid", start_color="F4B942", fgColor="F4B942")

    ws.merge_cells(
        start_row=total_row,
        start_column=count_col,
        end_row=total_row,
        end_column=pct_col,
    )
    for cell in (gt_count, gt_pct):
        cell.font = Font(bold=True, name="Arial")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", start_color="F4B942", fgColor="F4B942")

    for row_num in range(start_row, total_row + 1):
        for col_num in range(label_col, pct_col + 1):
            ws.cell(row=row_num, column=col_num).border = THIN_BORDER

    pct_col_letter = get_column_letter(pct_col)
    ws.conditional_formatting.add(
        f"{pct_col_letter}{start_row}:{pct_col_letter}{last_count_row}",
        PCT_COLOR_SCALE,
    )

    return total_row + 1


def _write_pivot_campaigns(
    ws,
    start_row: int,
    start_col: int,
    rows: list[dict[str, Any]],
    data_sheet_name: str = "CRM Output",
    filter_country: str | None = None,
    main_sheet_name: str | None = None,
) -> int:
    campaign_data: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for row in rows:
        campaign = str(row.get("Campaign", "") or "").strip() or "(No Campaign)"
        status = str(row.get("Status", "") or "").strip()
        if status:
            campaign_data[campaign][status] += 1

    campaigns_sorted = sorted(
        campaign_data.keys(),
        key=lambda campaign: (-sum(campaign_data[campaign].values()), campaign.lower()),
    )

    status_col = start_col
    count_col = start_col + 1
    pct_col = start_col + 2

    ref_sheet = main_sheet_name if (filter_country and main_sheet_name) else data_sheet_name
    status_letter = get_column_letter(OUTPUT_COLUMNS.index("Status") + 1)
    campaign_letter = get_column_letter(OUTPUT_COLUMNS.index("Campaign") + 1)
    status_range = f"'{ref_sheet}'!{status_letter}:{status_letter}"
    campaign_range = f"'{ref_sheet}'!{campaign_letter}:{campaign_letter}"
    count_col_letter = get_column_letter(count_col)

    if filter_country and main_sheet_name:
        country_letter = get_column_letter(OUTPUT_COLUMNS.index("Country") + 1)
        country_range = f"'{ref_sheet}'!{country_letter}:{country_letter}"
        safe_country = filter_country.replace('"', '""')

    n = len(STATUS_LIST)
    current_row = start_row
    campaign_total_addrs: list[str] = []
    pct_data_ranges: list[str] = []
    pct_col_letter = get_column_letter(pct_col)

    campaign_header_rows = 2

    for campaign in campaigns_sorted:
        header_start = current_row
        header_end = current_row + campaign_header_rows - 1
        header_cell = ws.cell(row=header_start, column=status_col, value=campaign)
        header_cell.font = Font(bold=True, name="Arial", size=12)
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
        header_cell.fill = PatternFill("solid", start_color="F4B942", fgColor="F4B942")
        ws.merge_cells(
            start_row=header_start,
            start_column=status_col,
            end_row=header_end,
            end_column=pct_col,
        )
        for row_num in range(header_start, header_end + 1):
            for col_num in range(status_col, pct_col + 1):
                ws.cell(row=row_num, column=col_num).border = THIN_BORDER

        first_count_row = header_end + 1
        last_count_row = first_count_row + n - 1
        campaign_total_row = last_count_row + 1
        campaign_total_addr = f"{count_col_letter}{campaign_total_row}"
        campaign_total_addrs.append(campaign_total_addr)
        pct_data_ranges.append(
            f"{pct_col_letter}{first_count_row}:{pct_col_letter}{last_count_row}"
        )

        safe_campaign = campaign.replace('"', '""')

        for i, status in enumerate(STATUS_LIST):
            row_i = first_count_row + i

            status_cell = ws.cell(row=row_i, column=status_col, value=status)
            status_cell.alignment = Alignment(horizontal="center", vertical="center")
            status_cell.font = Font(name="Arial")

            safe_status = status.replace('"', '""')
            if filter_country and main_sheet_name:
                count_formula = (
                    f'=COUNTIFS({campaign_range},"{safe_campaign}",'
                    f'{status_range},"{safe_status}",'
                    f'{country_range},"{safe_country}")'
                )
            else:
                count_formula = (
                    f'=COUNTIFS({campaign_range},"{safe_campaign}",'
                    f'{status_range},"{safe_status}")'
                )
            count_cell = ws.cell(row=row_i, column=count_col, value=count_formula)
            count_cell.alignment = Alignment(horizontal="center", vertical="center")
            count_cell.font = Font(name="Arial")

            count_cell_addr = f"{count_col_letter}{row_i}"
            pct_cell = ws.cell(
                row=row_i,
                column=pct_col,
                value=f"=IFERROR({count_cell_addr}/{campaign_total_addr},0)",
            )
            pct_cell.number_format = "0%"
            pct_cell.alignment = Alignment(horizontal="center", vertical="center")
            pct_cell.font = Font(name="Arial", bold=True)

        campaign_total_formula = (
            f"=SUM({count_col_letter}{first_count_row}:"
            f"{count_col_letter}{last_count_row})"
        )

        gt_label = ws.cell(row=campaign_total_row, column=status_col, value="Total")
        gt_label.font = Font(bold=True, name="Arial")
        gt_label.alignment = Alignment(horizontal="center", vertical="center")
        gt_label.fill = PatternFill("solid", start_color="F4B942", fgColor="F4B942")

        gt_count = ws.cell(
            row=campaign_total_row,
            column=count_col,
            value=campaign_total_formula,
        )
        gt_pct = ws.cell(row=campaign_total_row, column=pct_col, value="")
        ws.merge_cells(
            start_row=campaign_total_row,
            start_column=count_col,
            end_row=campaign_total_row,
            end_column=pct_col,
        )
        for cell in (gt_count, gt_pct):
            cell.font = Font(bold=True, name="Arial")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill("solid", start_color="F4B942", fgColor="F4B942")

        for row_num in range(first_count_row, campaign_total_row + 1):
            for col_num in range(status_col, pct_col + 1):
                ws.cell(row=row_num, column=col_num).border = THIN_BORDER

        current_row = campaign_total_row + 2

    overall_value: Any = "=" + "+".join(campaign_total_addrs) if campaign_total_addrs else 0
    for col_num, value in [
        (status_col, "Total"),
        (count_col, overall_value),
        (pct_col, ""),
    ]:
        cell = ws.cell(row=current_row, column=col_num, value=value)
        cell.font = Font(bold=True, name="Arial", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", start_color="BDD7EE", fgColor="BDD7EE")
        cell.border = THIN_BORDER

    for range_ref in pct_data_ranges:
        ws.conditional_formatting.add(range_ref, PCT_COLOR_SCALE)

    return current_row + 1


def _make_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)


def _make_font(hex_color: str, bold: bool = False) -> Font:
    return Font(color=hex_color, bold=bold, name="Arial")


def _apply_all_borders(worksheet, total_rows: int, total_cols: int) -> None:
    for row in worksheet.iter_rows(
        min_row=1,
        max_row=total_rows,
        min_col=1,
        max_col=total_cols,
    ):
        for cell in row:
            cell.border = THIN_BORDER


def write_output(
    rows: list[dict[str, Any]],
    output_file: Path,
    pivot_name: str,
) -> None:
    workbook = Workbook()
    ws_data = workbook.active
    ws_data.title = "CRM Output"

    status_col_idx = OUTPUT_COLUMNS.index("Status") + 1
    cb_col_idx = OUTPUT_COLUMNS.index("CB") + 1
    comments_col_idx = OUTPUT_COLUMNS.index("Comments") + 1

    ws_data.append(OUTPUT_COLUMNS)
    for cell in ws_data[1]:
        cell.font = Font(bold=True, name="Arial")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        data = [row.get(column, "") for column in OUTPUT_COLUMNS]

        cb_index = OUTPUT_COLUMNS.index("CB")
        if data[cb_index] is None:
            data[cb_index] = ""

        ws_data.append(data)
        excel_row = ws_data.max_row
        status_norm = normalize_status(row.get("Status", ""))

        status_cell = ws_data.cell(row=excel_row, column=status_col_idx)
        if NO_ANSWER_STATUS_RE.match(status_norm):
            fill_hex, font_hex = NO_ANSWER_COLORS
        else:
            fill_hex, font_hex = STATUS_COLORS.get(status_norm, (None, None))

        if fill_hex:
            status_cell.fill = _make_fill(fill_hex)
            status_cell.font = _make_font(font_hex)

        cb_cell = ws_data.cell(row=excel_row, column=cb_col_idx)
        if row.get("CB") is None:
            cb_cell.fill = _make_fill("000000")

        comments_cell = ws_data.cell(row=excel_row, column=comments_col_idx)
        comments_cell.alignment = Alignment(
            wrap_text=True,
            vertical="top",
            horizontal="center",
        )
        if row.get("_comments_yellow"):
            comments_cell.fill = _make_fill("FFFF00")

        for col_idx in range(1, len(OUTPUT_COLUMNS) + 1):
            cell = ws_data.cell(row=excel_row, column=col_idx)
            if col_idx != comments_col_idx:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            rgb = (
                cell.font.color.rgb
                if cell.font.color and cell.font.color.type == "rgb"
                else "000000"
            )
            cell.font = Font(name="Arial", color=rgb, bold=cell.font.bold)

    _apply_all_borders(ws_data, ws_data.max_row, len(OUTPUT_COLUMNS))

    last_col_letter = get_column_letter(len(OUTPUT_COLUMNS))
    last_data_row = ws_data.max_row
    table = Table(displayName="CRMOutput", ref=f"A1:{last_col_letter}{last_data_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws_data.add_table(table)

    for column_cells in ws_data.columns:
        header = column_cells[0].value
        max_length = max(len(str(cell.value or "")) for cell in column_cells[:100])
        width = min(max(max_length + 2, len(str(header or "")) + 2), 60)
        ws_data.column_dimensions[column_cells[0].column_letter].width = width

    pivot_start_row = last_data_row + 3
    current_row = pivot_start_row
    current_col = 1

    ws_data.column_dimensions["A"].width = max(
        ws_data.column_dimensions["A"].width,
        20,
    )
    ws_data.column_dimensions["B"].width = max(
        ws_data.column_dimensions["B"].width,
        32,
    )
    ws_data.column_dimensions["C"].width = max(
        ws_data.column_dimensions["C"].width,
        10,
    )
    ws_data.column_dimensions["D"].width = max(
        ws_data.column_dimensions["D"].width,
        10,
    )

    current_row = _write_pivot_status(
        ws_data,
        current_row,
        current_col,
        pivot_name,
        rows,
        data_sheet_name="CRM Output",
    )
    current_row += 1

    current_row = _write_pivot_call_attempts(
        ws_data,
        current_row,
        current_col,
        rows,
        data_sheet_name="CRM Output",
    )
    current_row += 3

    _write_pivot_campaigns(
        ws_data,
        current_row,
        current_col,
        rows,
        data_sheet_name="CRM Output",
    )

    output_file.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_file)


def build_output(
    powerbi_report: Path,
    crm_files: list[Path],
    platforms: list[str],
    pivot_name: str,
    output_file: Path,
    powerbi_sheet: str | None = None,
    crm_sheet: str | None = None,
) -> None:
    if len(crm_files) != len(platforms):
        raise ValueError("Each CRM file must have exactly one platform name.")

    powerbi_lookup = read_powerbi_lookup(powerbi_report, powerbi_sheet)
    all_rows: list[dict[str, Any]] = []

    for crm_file, platform in zip(crm_files, platforms):
        file_rows = read_crm_rows(crm_file, platform, powerbi_lookup, crm_sheet)
        all_rows.extend(file_rows)

    write_output(all_rows, output_file, pivot_name=pivot_name)


def prompt(message: str, allow_empty: bool = False) -> str:
    while True:
        value = input(message).strip()
        if value or allow_empty:
            return value
        print("  This field cannot be empty. Please try again.", file=sys.stderr)


def prompt_path(message: str, must_exist: bool = True) -> Path:
    while True:
        raw = prompt(message)
        path = Path(raw)
        if must_exist and not path.exists():
            print(
                f"  File not found: {path}. Please check the path and try again.",
                file=sys.stderr,
            )
            continue
        return path


def prompt_optional(message: str) -> str | None:
    value = input(message).strip()
    return value if value else None


def collect_inputs() -> dict[str, Any]:
    print("\n=== CRM + PowerBI Merge Tool ===\n")

    powerbi_report = prompt_path("PowerBI report file path: ")
    powerbi_sheet = prompt_optional(
        "  Sheet name (leave blank to use the active sheet): "
    )

    pivot_name = prompt("\nPivot table name (used as the label on the status pivot): ")

    crm_files: list[Path] = []
    platforms: list[str] = []

    print("\nEnter CRM files one at a time. Leave the path blank when done.")
    while True:
        index = len(crm_files) + 1
        crm_path_input = input(
            f"CRM file #{index} path (or press Enter to finish): "
        ).strip()

        if not crm_path_input:
            if not crm_files:
                print("  You must provide at least one CRM file.", file=sys.stderr)
                continue
            break

        crm_path = Path(crm_path_input)
        if not crm_path.exists():
            print(
                f"  File not found: {crm_path}. Please check the path and try again.",
                file=sys.stderr,
            )
            continue

        platform = prompt(f"  Platform name for '{crm_path.name}': ")
        crm_files.append(crm_path)
        platforms.append(platform)

    crm_sheet = prompt_optional(
        "\nCRM sheet name for all CRM files (leave blank for each file's active sheet): "
    )

    output_raw = input(
        "\nOutput file path (leave blank for 'crm_powerbi_output.xlsx'): "
    ).strip()
    output_file = Path(output_raw) if output_raw else Path("crm_powerbi_output.xlsx")

    return {
        "powerbi_report": powerbi_report,
        "powerbi_sheet": powerbi_sheet,
        "crm_files": crm_files,
        "platforms": platforms,
        "pivot_name": pivot_name,
        "crm_sheet": crm_sheet,
        "output_file": output_file,
    }


def main() -> int:
    try:
        inputs = collect_inputs()
        print("\nProcessing...")
        build_output(
            powerbi_report=inputs["powerbi_report"],
            crm_files=inputs["crm_files"],
            platforms=inputs["platforms"],
            pivot_name=inputs["pivot_name"],
            output_file=inputs["output_file"],
            powerbi_sheet=inputs["powerbi_sheet"],
            crm_sheet=inputs["crm_sheet"],
        )
    except KeyboardInterrupt:
        print("\nCancelled.", file=sys.stderr)
        return 1
    except Exception as exc:
        print(f"\nError: {exc}", file=sys.stderr)
        return 1

    print(f"\nDone! Output written to: {inputs['output_file']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
