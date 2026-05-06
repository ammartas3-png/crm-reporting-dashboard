"""Create a CRM output workbook enriched with PowerBI comments and call attempts.

The generated workbook contains:
  * one main report sheet for all CRM rows
  * status and call-attempt summary tables below the main report table
  * one sheet per country, each with the same report table filtered to that
    country and a country-named status summary below the table

The script interactively prompts for all required file paths and settings.
"""

from __future__ import annotations

import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


# ---------------------------------------------------------------------------
# Column definitions
# ---------------------------------------------------------------------------

CRM_COLUMNS = [
    "Customer Type",
    "ID",
    "Created",
    "Name",
    "Department",
    "Status",
    "Country",
    "Campaign",
    "Sub-Campaign",
    "Placement",
    "Assigned to",
]

POWERBI_COLUMNS = [
    "Account No",
    "Brand",
    "Last 10 Comments",
    "Voip Calls Attempts Cnt",
]

OUTPUT_COLUMNS = [
    "Platform",
    "Customer Type",
    "ID",
    "Created",
    "Name",
    "Department",
    "Status",
    "CB",
    "Country",
    "Campaign",
    "Sub-Campaign",
    "Placement",
    "Assigned to",
    "Comments",
    "Call Attempts",
]

DEFAULT_MAIN_SHEET_NAME = "Main Report"
DEFAULT_MAIN_STATUS_PIVOT_NAME = "M-Inhousemedia"


# ---------------------------------------------------------------------------
# Status colour map  (fill_hex, font_hex)
# ---------------------------------------------------------------------------

STATUS_COLORS: dict[str, tuple[str, str]] = {
    "call again": ("FFFF00", "000000"),
    "decline": ("FF0000", "FFFFFF"),
    "denied registration": ("ADD8E6", "000000"),
    "duplicate": ("4B0082", "FFFFFF"),
    "no interest": ("8B0000", "FFFFFF"),
    "no language": ("D3D3D3", "000000"),
    "no potential": ("FFB6C1", "000000"),
    "no potential - no bank account": ("0000FF", "000000"),
    "no potential - no documents": ("FFD580", "000000"),
    "not interested": ("8B0000", "FFFFFF"),
    "potential": ("90EE90", "000000"),
    "recall": ("404040", "FFFFFF"),
    "telemarketing": ("006400", "FFFFFF"),
    "under 18": ("FFA500", "000000"),
    "wrong number or email": ("D3D3D3", "000000"),
}

NO_ANSWER_STATUS_RE = re.compile(r"no answer ([1-5]|5\s*up)", re.IGNORECASE)
NO_ANSWER_COLORS = ("FF9999", "000000")

NO_COMMENT_STATUSES = {
    "dnc",
    "invalid country",
    "wrong number or email",
    "duplicate",
    "no potential - no documents",
    "test",
    "under 18",
    "no language",
}

COMMENT_RE = re.compile(r"\|\s*([^|;]*?)\s*;")

# Matches NA, NA VM, VM, DVM (standalone, whole comment)
NA_LIKE_PATTERN = re.compile(r"^(na(\s+vm)?|vm|dvm)$", re.IGNORECASE)

# Status words/phrases that should be silently dropped from comment text
STATUS_COMMENT_RE = re.compile(
    r"^("
    r"call again|decline|denied registration|duplicate"
    r"|no answer [1-5]|no answer 5\s*up"
    r"|no language|no potential|no potential\s*[-–]\s*no bank account"
    r"|no potential\s*[-–]\s*no documents"
    r"|not interested|no interest"
    r"|in progress"
    r"|potential|recall|under 18|wrong number or email"
    r")$",
    re.IGNORECASE,
)

# L1-L5 / V1-V5 prefix at the start of a comment - strip the tag, keep the rest
LV_PREFIX_RE = re.compile(r"^[lv][1-5]\s+", re.IGNORECASE)

INVALID_SHEET_CHARS_RE = re.compile(r"[\[\]:*?/\\]")
INVALID_TABLE_CHARS_RE = re.compile(r"[^A-Za-z0-9_]")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


# ---------------------------------------------------------------------------
# Normalisation helpers
# ---------------------------------------------------------------------------

def normalize_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


def normalize_status(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


def normalize_match_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"\s+", " ", str(value).strip()).casefold()


def clean_comment(value: str) -> str:
    return " ".join(value.split())


def display_text(value: Any, fallback: str = "") -> str:
    text = str(value or "").strip()
    return text if text else fallback


# ---------------------------------------------------------------------------
# Comment extraction & formatting
# ---------------------------------------------------------------------------

def extract_comments(last_10_comments: Any) -> list[str]:
    if last_10_comments is None:
        return []
    comments = [
        clean_comment(match.group(1))
        for match in COMMENT_RE.finditer(str(last_10_comments))
    ]
    comments = [comment for comment in comments if comment]
    comments.reverse()
    return comments


def _is_na_like(text: str) -> bool:
    return bool(NA_LIKE_PATTERN.match(text.strip()))


def _preprocess_comment(text: str) -> str | None:
    if text.lower().startswith("email"):
        return None
    if STATUS_COMMENT_RE.match(text.strip()):
        return None
    text = LV_PREFIX_RE.sub("", text).strip()
    return text if text else None


def format_comments(comments: list[str]) -> str:
    preprocessed: list[str] = []
    for comment in comments:
        result = _preprocess_comment(comment)
        if result is not None:
            preprocessed.append(result)

    merged: list[str] = []
    i = 0
    while i < len(preprocessed):
        if _is_na_like(preprocessed[i]):
            run = 1
            while i + run < len(preprocessed) and _is_na_like(preprocessed[i + run]):
                run += 1
            merged.append(f"NA VM x{run}" if run > 1 else "NA VM")
            i += run
        else:
            merged.append(preprocessed[i])
            i += 1

    return " // ".join(merged)


def comments_for_status(
    status: Any,
    matched_comments: list[str],
    found_in_powerbi: bool,
) -> tuple[str, bool]:
    normalized = normalize_status(status)

    if NO_ANSWER_STATUS_RE.match(normalized):
        return "NA VM", False

    if normalized in NO_COMMENT_STATUSES:
        return "", False

    if not found_in_powerbi:
        return "There was no comments on powerBI", True

    text = format_comments(matched_comments)
    if not text:
        return "There was no comments on powerBI", True

    return text, False


# ---------------------------------------------------------------------------
# Workbook input helpers
# ---------------------------------------------------------------------------

def worksheet_from_file(path: Path, sheet_name: str | None = None):
    workbook = load_workbook(path, data_only=True)
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"Sheet {sheet_name!r} not found in {path}. "
                f"Available sheets: {', '.join(workbook.sheetnames)}"
            )
        return workbook[sheet_name]
    return workbook.active


def header_indexes(
    headers: Iterable[Any],
    required_columns: list[str],
    path: Path,
) -> dict[str, int]:
    normalized_headers = {normalize_header(header): i for i, header in enumerate(headers)}
    missing = [
        column
        for column in required_columns
        if normalize_header(column) not in normalized_headers
    ]
    if missing:
        raise ValueError(f"{path} is missing required columns: {', '.join(missing)}")
    return {column: normalized_headers[normalize_header(column)] for column in required_columns}


# ---------------------------------------------------------------------------
# Core logic
# ---------------------------------------------------------------------------

def read_powerbi_lookup(
    powerbi_report: Path,
    sheet_name: str | None = None,
) -> dict[tuple[str, str], dict[str, Any]]:
    worksheet = worksheet_from_file(powerbi_report, sheet_name)
    rows = worksheet.iter_rows(values_only=True)

    try:
        headers = next(rows)
    except StopIteration as exc:
        raise ValueError(f"{powerbi_report} is empty") from exc

    indexes = header_indexes(headers, POWERBI_COLUMNS, powerbi_report)
    lookup: dict[tuple[str, str], dict[str, Any]] = {}

    for row in rows:
        account_no = row[indexes["Account No"]]
        brand_name = row[indexes["Brand"]]
        account_key = normalize_match_value(account_no)
        brand_key = normalize_match_value(brand_name)

        if not account_key or not brand_key:
            continue

        lookup[(account_key, brand_key)] = {
            "Comments": extract_comments(row[indexes["Last 10 Comments"]]),
            "Call Attempts": row[indexes["Voip Calls Attempts Cnt"]],
        }

    return lookup


def read_crm_rows(
    crm_file: Path,
    platform: str,
    powerbi_lookup: dict[tuple[str, str], dict[str, Any]],
    sheet_name: str | None = None,
) -> list[dict[str, Any]]:
    worksheet = worksheet_from_file(crm_file, sheet_name)
    rows = worksheet.iter_rows(values_only=True)

    try:
        headers = next(rows)
    except StopIteration as exc:
        raise ValueError(f"{crm_file} is empty") from exc

    indexes = header_indexes(headers, CRM_COLUMNS, crm_file)
    platform_key = normalize_match_value(platform)
    output_rows: list[dict[str, Any]] = []

    for row in rows:
        if all(value is None for value in row):
            continue

        crm_values = {column: row[indexes[column]] for column in CRM_COLUMNS}

        # Depositor override: force status to Telemarketing, no comments.
        customer_type_norm = normalize_status(crm_values.get("Customer Type", ""))
        if customer_type_norm == "depositor":
            crm_values["Status"] = "Telemarketing"
            output_rows.append(
                {
                    "Platform": platform,
                    **crm_values,
                    "CB": None,
                    "Comments": "",
                    "_comments_yellow": False,
                    "Call Attempts": "",
                }
            )
            continue

        account_key = normalize_match_value(crm_values["ID"])
        matched = powerbi_lookup.get((account_key, platform_key))
        found_in_powerbi = matched is not None
        matched_comments: list[str] = matched.get("Comments", []) if matched else []

        comment_text, use_yellow = comments_for_status(
            crm_values["Status"],
            matched_comments,
            found_in_powerbi,
        )

        status_norm = normalize_status(crm_values["Status"])
        cb_value = "" if status_norm == "call again" else None

        output_rows.append(
            {
                "Platform": platform,
                **crm_values,
                "CB": cb_value,
                "Comments": comment_text,
                "_comments_yellow": use_yellow,
                "Call Attempts": matched.get("Call Attempts", "") if matched else "",
            }
        )

    return output_rows


# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------

def _make_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)


def _make_font(hex_color: str, bold: bool = False) -> Font:
    return Font(color=hex_color, bold=bold, name="Arial")


def _apply_all_borders(worksheet, total_rows: int, total_cols: int) -> None:
    for row in worksheet.iter_rows(
        min_row=1,
        max_row=total_rows,
        min_col=1,
        max_col=total_cols,
    ):
        for cell in row:
            cell.border = THIN_BORDER


def _centered_pivot_start_col(total_report_cols: int) -> int:
    pivot_width = 4
    return max(1, ((total_report_cols - pivot_width) // 2) + 1)


def _apply_percentage_color_scale(ws, pct_ranges: list[str]) -> None:
    if not pct_ranges:
        return

    # Native Excel conditional formatting: red low values, yellow midpoint,
    # green high values. Percentages are stored as numbers so Excel can grade them.
    rule = ColorScaleRule(
        start_type="min",
        start_color="FF0000",
        mid_type="percentile",
        mid_value=50,
        mid_color="FFFF00",
        end_type="max",
        end_color="00B050",
    )
    for cell_range in pct_ranges:
        ws.conditional_formatting.add(cell_range, rule)


def _safe_sheet_title(raw_title: Any, existing_titles: set[str]) -> str:
    base = INVALID_SHEET_CHARS_RE.sub(" ", display_text(raw_title, "No Country"))
    base = re.sub(r"\s+", " ", base).strip("' ").strip() or "No Country"
    base = base[:31]

    title = base
    suffix = 2
    while title in existing_titles:
        suffix_text = f" {suffix}"
        title = f"{base[:31 - len(suffix_text)]}{suffix_text}"
        suffix += 1

    existing_titles.add(title)
    return title


def _safe_table_name(raw_name: str, used_names: set[str]) -> str:
    name = INVALID_TABLE_CHARS_RE.sub("_", raw_name)
    if not name or not name[0].isalpha():
        name = f"Table_{name}"
    name = name[:240]

    candidate = name
    suffix = 2
    while candidate in used_names:
        suffix_text = f"_{suffix}"
        candidate = f"{name[:240 - len(suffix_text)]}{suffix_text}"
        suffix += 1

    used_names.add(candidate)
    return candidate


# ---------------------------------------------------------------------------
# Summary table helpers
# ---------------------------------------------------------------------------

def _pct_value(pct: float) -> float:
    """Return an Excel percentage value, with display floored at 1%."""
    return max(0.01, min(1.0, pct))


def _write_status_summary(
    ws,
    start_row: int,
    start_col: int,
    summary_name: str,
    rows: list[dict[str, Any]],
) -> tuple[int, list[str]]:
    """
    Write a status summary table and return (next free row, percentage ranges).

    Layout:
      label       | status      | count | percentage
      summary name| Status A    | 10    | 50%
      (merged)    | Status B    | 10    | 50%
                  | Grand Total | 20    |
    """
    status_counts: dict[str, int] = defaultdict(int)
    for row in rows:
        status = display_text(row.get("Status"))
        if status:
            status_counts[status] += 1

    sorted_statuses = sorted(status_counts.items(), key=lambda item: item[0].lower())
    total = sum(status_counts.values())
    status_row_count = max(len(sorted_statuses), 1)

    label_col = start_col
    status_col = start_col + 1
    count_col = start_col + 2
    pct_col = start_col + 3

    name_cell = ws.cell(row=start_row, column=label_col, value=summary_name)
    name_cell.font = Font(bold=True, name="Arial", size=11)
    name_cell.alignment = Alignment(horizontal="center", vertical="center")
    name_cell.fill = PatternFill("solid", start_color="F4B942", fgColor="F4B942")

    if status_row_count > 1:
        ws.merge_cells(
            start_row=start_row,
            start_column=label_col,
            end_row=start_row + status_row_count - 1,
            end_column=label_col,
        )

    if sorted_statuses:
        for i, (status, count) in enumerate(sorted_statuses):
            row_i = start_row + i
            pct = count / total if total else 0

            status_cell = ws.cell(row=row_i, column=status_col, value=status)
            status_cell.alignment = Alignment(horizontal="center", vertical="center")
            status_cell.font = Font(name="Arial")

            count_cell = ws.cell(row=row_i, column=count_col, value=count)
            count_cell.alignment = Alignment(horizontal="center", vertical="center")
            count_cell.font = Font(name="Arial")

            pct_cell = ws.cell(row=row_i, column=pct_col, value=_pct_value(pct))
            pct_cell.alignment = Alignment(horizontal="center", vertical="center")
            pct_cell.number_format = "0%"
            pct_cell.font = Font(name="Arial", bold=True)
    else:
        empty_row = start_row
        for col, value in [(status_col, "No statuses"), (count_col, 0), (pct_col, None)]:
            cell = ws.cell(row=empty_row, column=col, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(name="Arial")

    total_row = start_row + status_row_count
    for col, value in [
        (label_col, ""),
        (status_col, "Grand Total"),
        (count_col, total),
        (pct_col, ""),
    ]:
        cell = ws.cell(row=total_row, column=col, value=value)
        cell.font = Font(bold=True, name="Arial")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", start_color="D9D9D9", fgColor="D9D9D9")

    for row_i in range(start_row, total_row + 1):
        for col_i in range(label_col, pct_col + 1):
            ws.cell(row=row_i, column=col_i).border = THIN_BORDER

    pct_ranges = []
    if sorted_statuses:
        pct_letter = get_column_letter(pct_col)
        pct_ranges.append(f"{pct_letter}{start_row}:{pct_letter}{start_row + len(sorted_statuses) - 1}")

    return total_row + 1, pct_ranges


def _write_call_attempts_summary(
    ws,
    start_row: int,
    start_col: int,
    rows: list[dict[str, Any]],
) -> tuple[int, list[str]]:
    """Write the Call Attempts summary table and return percentage ranges."""
    bucket_counts: dict[str, int] = defaultdict(int)
    for row in rows:
        raw = row.get("Call Attempts", "")
        try:
            value = int(float(str(raw))) if raw not in ("", None) else None
        except (TypeError, ValueError):
            value = None

        if value is None:
            continue
        if value <= 4:
            bucket_counts[str(value)] += 1
        else:
            bucket_counts["5+"] += 1

    bucket_order = ["1", "2", "3", "4", "5+"]
    total = sum(bucket_counts.values())

    label_col = start_col
    bucket_col = start_col + 1
    count_col = start_col + 2
    pct_col = start_col + 3

    name_cell = ws.cell(row=start_row, column=label_col, value="Call Attempts")
    name_cell.font = Font(bold=True, name="Arial", size=11)
    name_cell.alignment = Alignment(horizontal="center", vertical="center")
    name_cell.fill = PatternFill("solid", start_color="F4B942", fgColor="F4B942")
    ws.merge_cells(
        start_row=start_row,
        start_column=label_col,
        end_row=start_row + len(bucket_order) - 1,
        end_column=label_col,
    )

    for i, bucket in enumerate(bucket_order):
        row_i = start_row + i
        count = bucket_counts.get(bucket, 0)
        pct = count / total if total else 0

        bucket_cell = ws.cell(row=row_i, column=bucket_col, value=bucket)
        bucket_cell.alignment = Alignment(horizontal="center", vertical="center")
        bucket_cell.font = Font(name="Arial")

        count_cell = ws.cell(row=row_i, column=count_col, value=count)
        count_cell.alignment = Alignment(horizontal="center", vertical="center")
        count_cell.font = Font(name="Arial")

        pct_cell = ws.cell(row=row_i, column=pct_col, value=_pct_value(pct))
        pct_cell.alignment = Alignment(horizontal="center", vertical="center")
        pct_cell.number_format = "0%"
        pct_cell.font = Font(name="Arial", bold=True)

    total_row = start_row + len(bucket_order)
    for col, value in [
        (label_col, ""),
        (bucket_col, "Grand Total"),
        (count_col, total),
        (pct_col, ""),
    ]:
        cell = ws.cell(row=total_row, column=col, value=value)
        cell.font = Font(bold=True, name="Arial")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", start_color="D9D9D9", fgColor="D9D9D9")

    for row_i in range(start_row, total_row + 1):
        for col_i in range(label_col, pct_col + 1):
            ws.cell(row=row_i, column=col_i).border = THIN_BORDER

    pct_letter = get_column_letter(pct_col)
    pct_range = f"{pct_letter}{start_row}:{pct_letter}{start_row + len(bucket_order) - 1}"
    return total_row + 1, [pct_range]


# ---------------------------------------------------------------------------
# Output writer
# ---------------------------------------------------------------------------

def _write_report_table(
    ws,
    rows: list[dict[str, Any]],
    table_name: str,
) -> int:
    """Write the report rows and return the first row below the table."""
    status_col_idx = OUTPUT_COLUMNS.index("Status") + 1
    cb_col_idx = OUTPUT_COLUMNS.index("CB") + 1
    comments_col_idx = OUTPUT_COLUMNS.index("Comments") + 1

    ws.append(OUTPUT_COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True, name="Arial")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        data = [row.get(column, "") for column in OUTPUT_COLUMNS]

        cb_index = OUTPUT_COLUMNS.index("CB")
        if data[cb_index] is None:
            data[cb_index] = ""

        ws.append(data)
        excel_row = ws.max_row
        status_norm = normalize_status(row.get("Status", ""))

        status_cell = ws.cell(row=excel_row, column=status_col_idx)
        if NO_ANSWER_STATUS_RE.match(status_norm):
            fill_hex, font_hex = NO_ANSWER_COLORS
        else:
            fill_hex, font_hex = STATUS_COLORS.get(status_norm, (None, None))

        if fill_hex:
            status_cell.fill = _make_fill(fill_hex)
            status_cell.font = _make_font(font_hex)

        cb_cell = ws.cell(row=excel_row, column=cb_col_idx)
        if row.get("CB") is None:
            cb_cell.fill = _make_fill("000000")

        comments_cell = ws.cell(row=excel_row, column=comments_col_idx)
        comments_cell.alignment = Alignment(
            wrap_text=True,
            vertical="top",
            horizontal="center",
        )
        if row.get("_comments_yellow"):
            comments_cell.fill = _make_fill("FFFF00")

        for col_idx in range(1, len(OUTPUT_COLUMNS) + 1):
            cell = ws.cell(row=excel_row, column=col_idx)
            if col_idx != comments_col_idx:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            rgb = (
                cell.font.color.rgb
                if cell.font.color and cell.font.color.type == "rgb"
                else "000000"
            )
            cell.font = Font(name="Arial", color=rgb, bold=cell.font.bold)

    _apply_all_borders(ws, ws.max_row, len(OUTPUT_COLUMNS))

    last_col_letter = get_column_letter(len(OUTPUT_COLUMNS))
    last_row = ws.max_row
    table = Table(displayName=table_name, ref=f"A1:{last_col_letter}{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    for column_cells in ws.columns:
        header = column_cells[0].value
        max_length = max(len(str(cell.value or "")) for cell in column_cells[:100])
        width = min(max(max_length + 2, len(str(header or "")) + 2), 60)
        ws.column_dimensions[column_cells[0].column_letter].width = width

    return last_row + 1


def _write_main_summaries(
    ws,
    start_row: int,
    rows: list[dict[str, Any]],
    main_status_pivot_name: str,
) -> None:
    pivot_col = _centered_pivot_start_col(len(OUTPUT_COLUMNS))
    current_row = start_row + 2
    pct_ranges: list[str] = []

    current_row, ranges = _write_status_summary(
        ws,
        current_row,
        pivot_col,
        main_status_pivot_name,
        rows,
    )
    pct_ranges.extend(ranges)

    current_row += 1
    _, ranges = _write_call_attempts_summary(ws, current_row, pivot_col, rows)
    pct_ranges.extend(ranges)

    _apply_percentage_color_scale(ws, pct_ranges)


def _write_country_summary(
    ws,
    start_row: int,
    rows: list[dict[str, Any]],
    country: str,
) -> None:
    pivot_col = _centered_pivot_start_col(len(OUTPUT_COLUMNS))
    current_row = start_row + 2
    _, pct_ranges = _write_status_summary(ws, current_row, pivot_col, country, rows)
    _apply_percentage_color_scale(ws, pct_ranges)


def write_output(
    rows: list[dict[str, Any]],
    output_file: Path,
    main_status_pivot_name: str = DEFAULT_MAIN_STATUS_PIVOT_NAME,
) -> None:
    workbook = Workbook()
    used_table_names: set[str] = set()

    ws_main = workbook.active
    ws_main.title = DEFAULT_MAIN_SHEET_NAME
    table_name = _safe_table_name("MainReport", used_table_names)
    next_row = _write_report_table(ws_main, rows, table_name)
    _write_main_summaries(ws_main, next_row, rows, main_status_pivot_name)

    country_groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        country = display_text(row.get("Country"), "No Country")
        country_groups[country].append(row)

    existing_titles = set(workbook.sheetnames)
    for country in sorted(country_groups, key=str.lower):
        sheet_title = _safe_sheet_title(country, existing_titles)
        ws_country = workbook.create_sheet(title=sheet_title)
        table_name = _safe_table_name(f"{sheet_title}_Report", used_table_names)
        country_rows = country_groups[country]
        next_row = _write_report_table(ws_country, country_rows, table_name)
        _write_country_summary(ws_country, next_row, country_rows, country)

    output_file.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_file)


# ---------------------------------------------------------------------------
# Build orchestrator
# ---------------------------------------------------------------------------

def build_output(
    powerbi_report: Path,
    crm_files: list[Path],
    platforms: list[str],
    output_file: Path,
    powerbi_sheet: str | None = None,
    crm_sheet: str | None = None,
    main_status_pivot_name: str = DEFAULT_MAIN_STATUS_PIVOT_NAME,
) -> None:
    if len(crm_files) != len(platforms):
        raise ValueError("Each CRM file must have exactly one platform name.")

    powerbi_lookup = read_powerbi_lookup(powerbi_report, powerbi_sheet)
    all_rows: list[dict[str, Any]] = []

    for crm_file, platform in zip(crm_files, platforms):
        file_rows = read_crm_rows(crm_file, platform, powerbi_lookup, crm_sheet)
        all_rows.extend(file_rows)

    write_output(all_rows, output_file, main_status_pivot_name=main_status_pivot_name)


# ---------------------------------------------------------------------------
# Interactive prompts
# ---------------------------------------------------------------------------

def prompt(message: str, allow_empty: bool = False) -> str:
    while True:
        value = input(message).strip()
        if value or allow_empty:
            return value
        print("  This field cannot be empty. Please try again.", file=sys.stderr)


def prompt_path(message: str, must_exist: bool = True) -> Path:
    while True:
        raw = prompt(message)
        path = Path(raw)
        if must_exist and not path.exists():
            print(f"  File not found: {path}. Please check the path and try again.", file=sys.stderr)
            continue
        return path


def prompt_optional(message: str) -> str | None:
    value = input(message).strip()
    return value if value else None


def collect_inputs() -> dict[str, Any]:
    print("\n=== CRM + PowerBI Country Report Tool ===\n")

    powerbi_report = prompt_path("PowerBI report file path: ")
    powerbi_sheet = prompt_optional("  Sheet name (leave blank to use the active sheet): ")

    crm_files: list[Path] = []
    platforms: list[str] = []

    print("\nEnter CRM files one at a time. Leave the path blank when done.")
    while True:
        index = len(crm_files) + 1
        crm_path_input = input(f"CRM file #{index} path (or press Enter to finish): ").strip()

        if not crm_path_input:
            if not crm_files:
                print("  You must provide at least one CRM file.", file=sys.stderr)
                continue
            break

        crm_path = Path(crm_path_input)
        if not crm_path.exists():
            print(f"  File not found: {crm_path}. Please check the path and try again.", file=sys.stderr)
            continue

        platform = prompt(f"  Platform name for '{crm_path.name}' / PowerBI Brand match: ")
        crm_files.append(crm_path)
        platforms.append(platform)

    crm_sheet = prompt_optional("\nCRM sheet name for all CRM files (leave blank for each file's active sheet): ")
    main_status_pivot_name = (
        prompt_optional(
            f"Main status pivot name (leave blank for '{DEFAULT_MAIN_STATUS_PIVOT_NAME}'): "
        )
        or DEFAULT_MAIN_STATUS_PIVOT_NAME
    )

    output_raw = input("\nOutput file path (leave blank for 'crm_country_report_output.xlsx'): ").strip()
    output_file = Path(output_raw) if output_raw else Path("crm_country_report_output.xlsx")

    return {
        "powerbi_report": powerbi_report,
        "powerbi_sheet": powerbi_sheet,
        "crm_files": crm_files,
        "platforms": platforms,
        "crm_sheet": crm_sheet,
        "main_status_pivot_name": main_status_pivot_name,
        "output_file": output_file,
    }


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> int:
    try:
        inputs = collect_inputs()
        print("\nProcessing...")
        build_output(
            powerbi_report=inputs["powerbi_report"],
            crm_files=inputs["crm_files"],
            platforms=inputs["platforms"],
            output_file=inputs["output_file"],
            powerbi_sheet=inputs["powerbi_sheet"],
            crm_sheet=inputs["crm_sheet"],
            main_status_pivot_name=inputs["main_status_pivot_name"],
        )
    except KeyboardInterrupt:
        print("\nCancelled.", file=sys.stderr)
        return 1
    except Exception as exc:  # noqa: BLE001
        print(f"\nError: {exc}", file=sys.stderr)
        return 1

    print(f"\nDone! Output written to: {inputs['output_file']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
