"""Create a CRM output workbook enriched with PowerBI comments and call attempts.

The command line interface accepts one PowerBI report and one or more CRM Excel
files. For each CRM file, it asks for the platform name unless it was provided
with --platform.
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font


CRM_COLUMNS = [
    "Customer Type",
    "ID",
    "Created",
    "Name",
    "Department",
    "Status",
    "Country",
    "Assigned to",
]

POWERBI_COLUMNS = [
    "Account No",
    "Brand name",
    "Last 10 Comments",
    "Voip Calls Attempts Cnt",
]

OUTPUT_COLUMNS = [
    "Platform",
    *CRM_COLUMNS,
    "Comments",
    "Call Attempts",
]

NO_COMMENT_STATUSES = {
    "dnc",
    "invalid country",
    "wrong number or email",
    "duplicate",
    "no potential - no documents",
    "test",
    "under 18",
    "no language",
}

NO_ANSWER_STATUS_RE = re.compile(r"^no\s*answer\s*([1-5])$", re.IGNORECASE)

# Captures the final pipe-delimited field before a semicolon:
# "date | agent | comment;" -> "comment".
COMMENT_RE = re.compile(r"\|\s*([^|;]*?)\s*;")


def normalize_header(value: Any) -> str:
    """Normalize worksheet headers for forgiving column matching."""
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


def normalize_status(value: Any) -> str:
    """Normalize CRM statuses for business-rule comparisons."""
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


def normalize_match_value(value: Any) -> str:
    """Normalize IDs and platform/brand values used for matching rows."""
    if value is None:
        return ""

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    return re.sub(r"\s+", " ", str(value).strip()).casefold()


def clean_comment(value: str) -> str:
    """Remove wrapping whitespace inside one parsed comment."""
    return " ".join(value.split())


def extract_comments(last_10_comments: Any) -> str:
    """Extract comments from PowerBI text and return them from bottom to top."""
    if last_10_comments is None:
        return ""

    comments = [
        clean_comment(match.group(1))
        for match in COMMENT_RE.finditer(str(last_10_comments))
    ]
    comments = [comment for comment in comments if comment]
    comments.reverse()
    return "\n".join(comments)


def comments_for_status(status: Any, matched_comments: str) -> str:
    """Apply CRM status rules to the comments output field."""
    normalized = normalize_status(status)

    if NO_ANSWER_STATUS_RE.match(normalized):
        return "NA VM"

    if normalized in NO_COMMENT_STATUSES:
        return ""

    return matched_comments


def worksheet_from_file(path: Path, sheet_name: str | None = None):
    workbook = load_workbook(path, data_only=True)
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"Sheet {sheet_name!r} was not found in {path}. "
                f"Available sheets: {', '.join(workbook.sheetnames)}"
            )
        return workbook[sheet_name]
    return workbook.active


def header_indexes(headers: Iterable[Any], required_columns: list[str], path: Path) -> dict[str, int]:
    normalized_headers = {
        normalize_header(header): index for index, header in enumerate(headers)
    }

    missing = [
        column
        for column in required_columns
        if normalize_header(column) not in normalized_headers
    ]
    if missing:
        raise ValueError(f"{path} is missing required columns: {', '.join(missing)}")

    return {
        column: normalized_headers[normalize_header(column)]
        for column in required_columns
    }


def read_powerbi_lookup(
    powerbi_report: Path,
    sheet_name: str | None = None,
) -> dict[tuple[str, str], dict[str, Any]]:
    """Read PowerBI rows keyed by (Account No, Brand name)."""
    worksheet = worksheet_from_file(powerbi_report, sheet_name)
    rows = worksheet.iter_rows(values_only=True)

    try:
        headers = next(rows)
    except StopIteration as exc:
        raise ValueError(f"{powerbi_report} is empty") from exc

    indexes = header_indexes(headers, POWERBI_COLUMNS, powerbi_report)
    lookup: dict[tuple[str, str], dict[str, Any]] = {}

    for row in rows:
        account_no = row[indexes["Account No"]]
        brand_name = row[indexes["Brand name"]]
        account_key = normalize_match_value(account_no)
        brand_key = normalize_match_value(brand_name)

        if not account_key or not brand_key:
            continue

        lookup[(account_key, brand_key)] = {
            "Comments": extract_comments(row[indexes["Last 10 Comments"]]),
            "Call Attempts": row[indexes["Voip Calls Attempts Cnt"]],
        }

    return lookup


def read_crm_rows(
    crm_file: Path,
    platform: str,
    powerbi_lookup: dict[tuple[str, str], dict[str, Any]],
    sheet_name: str | None = None,
) -> list[dict[str, Any]]:
    """Read one CRM workbook and return normalized output rows."""
    worksheet = worksheet_from_file(crm_file, sheet_name)
    rows = worksheet.iter_rows(values_only=True)

    try:
        headers = next(rows)
    except StopIteration as exc:
        raise ValueError(f"{crm_file} is empty") from exc

    indexes = header_indexes(headers, CRM_COLUMNS, crm_file)
    platform_key = normalize_match_value(platform)
    output_rows: list[dict[str, Any]] = []

    for row in rows:
        if all(value is None for value in row):
            continue

        crm_values = {column: row[indexes[column]] for column in CRM_COLUMNS}
        account_key = normalize_match_value(crm_values["ID"])
        matched = powerbi_lookup.get((account_key, platform_key), {})

        output_rows.append(
            {
                "Platform": platform,
                **crm_values,
                "Comments": comments_for_status(
                    crm_values["Status"],
                    str(matched.get("Comments", "") or ""),
                ),
                "Call Attempts": matched.get("Call Attempts", ""),
            }
        )

    return output_rows


def write_output(rows: list[dict[str, Any]], output_file: Path) -> None:
    """Write output rows to a new Excel workbook."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "CRM Output"

    worksheet.append(OUTPUT_COLUMNS)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for row in rows:
        worksheet.append([row.get(column, "") for column in OUTPUT_COLUMNS])

    for column_cells in worksheet.columns:
        header = column_cells[0].value
        max_length = max(len(str(cell.value or "")) for cell in column_cells[:100])
        width = min(max(max_length + 2, len(str(header)) + 2), 60)
        worksheet.column_dimensions[column_cells[0].column_letter].width = width

    comments_column = OUTPUT_COLUMNS.index("Comments") + 1
    for cell in worksheet.iter_cols(
        min_col=comments_column,
        max_col=comments_column,
        min_row=2,
    ):
        for comment_cell in cell:
            comment_cell.alignment = Alignment(wrap_text=True, vertical="top")

    output_file.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_file)


def build_output(
    powerbi_report: Path,
    crm_files: list[Path],
    platforms: list[str],
    output_file: Path,
    powerbi_sheet: str | None = None,
    crm_sheet: str | None = None,
) -> None:
    """Build the requested output workbook."""
    if len(crm_files) != len(platforms):
        raise ValueError("Each CRM file must have exactly one platform name.")

    powerbi_lookup = read_powerbi_lookup(powerbi_report, powerbi_sheet)
    output_rows: list[dict[str, Any]] = []

    for crm_file, platform in zip(crm_files, platforms):
        output_rows.extend(read_crm_rows(crm_file, platform, powerbi_lookup, crm_sheet))

    write_output(output_rows, output_file)


def collect_platforms(crm_files: list[Path], provided_platforms: list[str]) -> list[str]:
    """Prompt for any CRM platforms not supplied on the command line."""
    if len(provided_platforms) > len(crm_files):
        raise ValueError("More --platform values were provided than CRM files.")

    platforms = list(provided_platforms)
    for crm_file in crm_files[len(platforms) :]:
        while True:
            platform = input(f"Platform for CRM file '{crm_file}': ").strip()
            if platform:
                platforms.append(platform)
                break
            print("Platform cannot be empty.", file=sys.stderr)

    return platforms


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Merge CRM Excel files with a PowerBI report and write the requested "
            "CRM output workbook."
        )
    )
    parser.add_argument(
        "--powerbi-report",
        required=True,
        type=Path,
        help="Path to the PowerBI Excel report.",
    )
    parser.add_argument(
        "--crm-files",
        required=True,
        nargs="+",
        type=Path,
        help="One or more CRM Excel files.",
    )
    parser.add_argument(
        "--output",
        default=Path("crm_powerbi_output.xlsx"),
        type=Path,
        help="Output Excel file path. Defaults to crm_powerbi_output.xlsx.",
    )
    parser.add_argument(
        "--platform",
        action="append",
        default=[],
        help=(
            "Platform for a CRM file. Repeat this option in the same order as "
            "--crm-files to skip interactive prompts."
        ),
    )
    parser.add_argument(
        "--powerbi-sheet",
        help="Optional PowerBI worksheet name. Defaults to the active sheet.",
    )
    parser.add_argument(
        "--crm-sheet",
        help="Optional CRM worksheet name used for every CRM file. Defaults to each active sheet.",
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)

    try:
        platforms = collect_platforms(args.crm_files, args.platform)
        build_output(
            args.powerbi_report,
            args.crm_files,
            platforms,
            args.output,
            args.powerbi_sheet,
            args.crm_sheet,
        )
    except Exception as exc:  # noqa: BLE001 - CLI should print concise errors.
        print(f"Error: {exc}", file=sys.stderr)
        return 1

    print(f"Wrote {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
