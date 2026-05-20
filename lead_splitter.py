from __future__ import annotations

import os
import re
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

NAME_FIXES = {
    "Ahmed Meh": "Ahmed Me",
    "Ali D": "Ali Da",
    "Fekraoui Ra": "FekRaraoui Ra",
    "Jehad Ab": "Jehad Ba",
    "Muhammed K": "Muhammed Ke",
    "Rissa RissaZa": "Rissa Za",
    "Nadia Ro": "Nadia R",
    "Aya C": "Aya Ca",
    "Ismail Alt": "Ismail Al",
    "Abrar Os": "Abrar O",
    "Ahmad Al": "Ahmad A",
    "Ali Nad": "Ali Na",
    "Ali Sal": "Ali Sa",
    "Ana Ca": "Ana C",
    "Anas Ab": "Anas B",
    "Astrolan Nn": "Astrolan No",
    "Bruno Es": "Bruno E",
    "Bruno EE": "Bruno E",
    "Camera Al": "Camera Ai",
    "Cecilia Ot": "Cecilia Ro",
    "Mehmet AkAE": "Mehmet Ak",
    "Melis SAE": "Melis Su",
    "Meric CeAE": "Meric Ce",
    "Elauterio Lima": "Elauterio Li",
    "Toualibi Ay": "Ben Ay",
    "Toulaibi Ay": "Ben Ay",
    "Manal ManalYa": "Manal Ya",
    "Pyae H": "Pyae He",
    "Mehmet K": "Mehmet Ki",
    "Melis Su": "Melis S",
    "Ali Yi": "Ali Yi",
    "Aristides QuAr": "Aristides Qu",
    "Ismai Al": "Ismail Al",
    "Muhammad MuhammadZa": "MuhammadZa",
    "Said .Al": "Said Al",
    "Manal Manal.Ya": "Manal Ya",
    "Tania Wib": "Tania Wi",
    "Syahmi Ra": "Syahmi Rah",
    "Rahma Far": "Rahma Fa",
    "Dilan. Ka": "Dilan Ka",
}

MY_EXEMPT_AGENTS = {
    "Shasha We",
    "Hatim La",
    "Said Al",
    "Ahmad Ma",
    "Nusaiba Ar",
    "Joel Go",
    "Heela An",
    "Ceydanur Em",
    "Kenneth Ba",
    "Abass Ad",
    "Joy Ot",
    "Esra Av",
    "Medisa Ta",
    "Selay Gu",
    "Didar Gu",
    "Taylan Bo",
    "Yahya Ze",
    "Mehmet Ki",
    "David Le",
    "Oluwasgun Oy",
    "Mostafa Va",
    "Suayib Mo",
    "Mohammed Ba",
    "Pyae He",
    "Aymane Ab",
    "Leyla Go",
    "Muhammad Ibr",
    "Ghada Aa",
    "Abrar Os",
    "Khadija A",
}

GCC_COUNTRIES = {
    "Saudi Arabia",
    "United Arab Emirates",
    "Qatar",
    "Kuwait",
    "Bahrain",
    "Oman",
}


def make_border(color: str = "D0D0D0") -> Border:
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def clean_agent_name(name):
    if not isinstance(name, str):
        return name
    name = re.sub(r"\s*\([^)]*\)\s*$", "", name).strip()
    name = re.sub(r"(CY|AE|IN)$", "", name).strip()
    return name


def is_pool_agent(name) -> bool:
    if not isinstance(name, str):
        return False
    clean = name.strip()
    return clean.startswith("BI pool") or clean.startswith("Pool")


def fix_name(name):
    return NAME_FIXES.get(name, name)


def get_desk2(desk) -> str:
    if not isinstance(desk, str):
        return str(desk)
    parts = desk.split("-")
    return parts[1] if len(parts) >= 2 else desk


def get_office(desk) -> str:
    if not isinstance(desk, str) or len(desk) < 2:
        return ""
    return str(desk)[:2].upper()


def build_pivot(wb, df, n_col, o_col, b_col, c_col, i_col) -> None:
    ws = wb.create_sheet("Pivot")

    df = df.copy()
    df["_DESK2"] = df[b_col].apply(get_desk2)
    df["_N1"] = df[n_col].apply(lambda x: 1 if str(x).strip() == "1" else 0)
    df["_O1"] = df[o_col].apply(lambda x: 1 if str(x).strip() == "1" else 0)

    agg = df.groupby(["_DESK2", i_col, c_col], sort=True).agg(
        Assigned=("_N1", "sum"), FTD=("_O1", "sum")
    ).reset_index()

    fill_header = PatternFill("solid", start_color="17375E", end_color="17375E")
    fill_desk_total = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
    fill_country_total = PatternFill("solid", start_color="DEEAF1", end_color="DEEAF1")
    fill_alt = PatternFill("solid", start_color="F5FBFF", end_color="F5FBFF")
    font_header = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    font_bold = Font(bold=True, color="000000", name="Arial", size=10)
    font_norm = Font(bold=False, color="000000", name="Arial", size=10)
    border = make_border("C8C8C8")

    def write_row(row_i, desk_val, country_val, agent_val, assigned, ftd, font, fill):
        cr = ftd / assigned if assigned > 0 else 0
        values = [desk_val, country_val, agent_val, assigned, ftd, cr]
        for col_i, val in enumerate(values, 1):
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.font = font
            cell.fill = fill
            cell.border = border
            if col_i <= 3:
                cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if col_i == 6:
                cell.number_format = "0%"
        ws.row_dimensions[row_i].height = 16

    for col_i, header in enumerate(
        ["Desk", "Country", "Agent", "Assigned to", "FTD Count", "CR"], 1
    ):
        cell = ws.cell(row=1, column=col_i, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = make_border("0D2340")
    ws.row_dimensions[1].height = 22

    cur_row = 2
    for desk2, desk_df in agg.groupby("_DESK2", sort=True):
        first_desk = True
        country_totals = desk_df.groupby(i_col)["Assigned"].sum().sort_values(ascending=False)
        country_order = country_totals.index.tolist()

        for country in country_order:
            country_df = desk_df[desk_df[i_col] == country]
            first_country = True
            alt = 0

            for _, row in country_df.iterrows():
                desk_label = desk2 if first_desk else ""
                country_label = country if first_country else ""
                fill = fill_alt if alt % 2 == 0 else PatternFill(fill_type=None)
                write_row(
                    cur_row,
                    desk_label,
                    country_label,
                    row[c_col],
                    int(row["Assigned"]),
                    int(row["FTD"]),
                    font_norm,
                    fill,
                )
                first_desk = False
                first_country = False
                alt += 1
                cur_row += 1

            country_assigned = int(country_df["Assigned"].sum())
            country_ftd = int(country_df["FTD"].sum())
            write_row(
                cur_row,
                "",
                f"{country} Total",
                "",
                country_assigned,
                country_ftd,
                font_bold,
                fill_country_total,
            )
            cur_row += 1

        desk_assigned = int(desk_df["Assigned"].sum())
        desk_ftd = int(desk_df["FTD"].sum())
        write_row(
            cur_row,
            f"{desk2} Total",
            "",
            "",
            desk_assigned,
            desk_ftd,
            font_bold,
            fill_desk_total,
        )
        cur_row += 1

    widths = {"A": 16, "B": 28, "C": 24, "D": 14, "E": 12, "F": 10}
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = "D2"


AFF_BORDER = make_border("C8C8C8")
AFF_BORDER_DARK = make_border("0D2340")
AFF_FILL_HEADER = PatternFill("solid", start_color="17375E", end_color="17375E")
AFF_FILL_CAMP = PatternFill("solid", start_color="2E75B6", end_color="2E75B6")
AFF_FILL_OFFICE = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
AFF_FILL_ALT = PatternFill("solid", start_color="F5FBFF", end_color="F5FBFF")
AFF_FILL_NONE = PatternFill(fill_type=None)
AFF_FILL_GRAND = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
AFF_FONT_HDR = Font(bold=True, color="FFFFFF", name="Arial", size=10)
AFF_FONT_BOLD = Font(bold=True, name="Arial", size=10)
AFF_FONT_BOLD_W = Font(bold=True, color="FFFFFF", name="Arial", size=10)
AFF_FONT_NORM = Font(bold=False, name="Arial", size=10)


def _cr(leads, ftd) -> float:
    return ftd / leads if leads > 0 else 0


def _aff_write_headers(ws, col_offset, headers) -> None:
    for j, header in enumerate(headers):
        cell = ws.cell(row=1, column=col_offset + j + 1, value=header)
        cell.font = AFF_FONT_HDR
        cell.fill = AFF_FILL_HEADER
        cell.border = AFF_BORDER_DARK
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20


def _aff_write_row(ws, row_i, col_offset, vals, font, fill, skip_fill=0) -> None:
    for j, val in enumerate(vals):
        cell = ws.cell(row=row_i, column=col_offset + j + 1, value=val)
        n_cols = len(vals)
        pct_col = n_cols - 1
        num_start = n_cols - 3
        if j < skip_fill:
            cell.font = AFF_FONT_NORM
            cell.fill = AFF_FILL_NONE
        else:
            cell.font = font
            cell.fill = fill
        cell.border = AFF_BORDER
        if j >= num_start:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        if j == pct_col:
            cell.number_format = "0%"
    ws.row_dimensions[row_i].height = 15


def _write_standard_table(ws, data_df, col_offset, country_label, campaign_col, status_col) -> None:
    _aff_write_headers(
        ws,
        col_offset,
        ["Country", "Office", "Campaign", "Status", "LEADS", "FTD", "CR%"],
    )
    row_i = 2
    first_country = True

    for office, office_df in data_df.groupby("_OFFICE", sort=True):
        first_office = True
        for campaign, campaign_df in office_df.groupby(campaign_col, sort=True):
            first_campaign = True
            alt = 0

            status_agg = (
                campaign_df.groupby(status_col, sort=True)
                .agg(Leads=("_N1", "sum"), FTD=("_O1", "sum"))
                .reset_index()
            )

            for _, status_row in status_agg.iterrows():
                leads = int(status_row["Leads"])
                ftd = int(status_row["FTD"])
                vals = [
                    country_label if first_country else "",
                    office if first_office else "",
                    campaign if first_campaign else "",
                    status_row[status_col],
                    leads,
                    ftd,
                    _cr(leads, ftd),
                ]
                _aff_write_row(
                    ws,
                    row_i,
                    col_offset,
                    vals,
                    AFF_FONT_NORM,
                    AFF_FILL_ALT if alt % 2 == 0 else AFF_FILL_NONE,
                )
                first_country = False
                first_office = False
                first_campaign = False
                alt += 1
                row_i += 1

            campaign_leads = int(campaign_df["_N1"].sum())
            campaign_ftd = int(campaign_df["_O1"].sum())
            _aff_write_row(
                ws,
                row_i,
                col_offset,
                ["", "", f"{campaign} Total", "", campaign_leads, campaign_ftd, _cr(campaign_leads, campaign_ftd)],
                AFF_FONT_BOLD_W,
                AFF_FILL_CAMP,
                skip_fill=2,
            )
            row_i += 1

        office_leads = int(office_df["_N1"].sum())
        office_ftd = int(office_df["_O1"].sum())
        _aff_write_row(
            ws,
            row_i,
            col_offset,
            ["", f"{office} Total", "", "", office_leads, office_ftd, _cr(office_leads, office_ftd)],
            AFF_FONT_BOLD,
            AFF_FILL_OFFICE,
        )
        row_i += 1

    grand_leads = int(data_df["_N1"].sum())
    grand_ftd = int(data_df["_O1"].sum())
    _aff_write_row(
        ws,
        row_i,
        col_offset,
        [f"{country_label} Total", "", "", "", grand_leads, grand_ftd, _cr(grand_leads, grand_ftd)],
        AFF_FONT_BOLD_W,
        AFF_FILL_GRAND,
    )


def _write_gcc_table(ws, data_df, col_offset, campaign_col, country_col, status_col) -> None:
    _aff_write_headers(
        ws,
        col_offset,
        ["Country", "Office", "Campaign", "Country", "Status", "LEADS", "FTD", "CR%"],
    )
    row_i = 2
    first_region = True

    for office, office_df in data_df.groupby("_OFFICE", sort=True):
        first_office = True
        for campaign, campaign_df in office_df.groupby(campaign_col, sort=True):
            first_campaign = True

            for country, country_df in campaign_df.groupby(country_col, sort=True):
                first_country = True
                alt = 0

                status_agg = (
                    country_df.groupby(status_col, sort=True)
                    .agg(Leads=("_N1", "sum"), FTD=("_O1", "sum"))
                    .reset_index()
                )

                for _, status_row in status_agg.iterrows():
                    leads = int(status_row["Leads"])
                    ftd = int(status_row["FTD"])
                    vals = [
                        "GCC EN" if first_region else "",
                        office if first_office else "",
                        campaign if first_campaign else "",
                        country if first_country else "",
                        status_row[status_col],
                        leads,
                        ftd,
                        _cr(leads, ftd),
                    ]
                    _aff_write_row(
                        ws,
                        row_i,
                        col_offset,
                        vals,
                        AFF_FONT_NORM,
                        AFF_FILL_ALT if alt % 2 == 0 else AFF_FILL_NONE,
                    )
                    first_region = False
                    first_office = False
                    first_campaign = False
                    first_country = False
                    alt += 1
                    row_i += 1

            campaign_leads = int(campaign_df["_N1"].sum())
            campaign_ftd = int(campaign_df["_O1"].sum())
            _aff_write_row(
                ws,
                row_i,
                col_offset,
                ["", "", f"{campaign} Total", "", "", campaign_leads, campaign_ftd, _cr(campaign_leads, campaign_ftd)],
                AFF_FONT_BOLD_W,
                AFF_FILL_CAMP,
                skip_fill=2,
            )
            row_i += 1

        office_leads = int(office_df["_N1"].sum())
        office_ftd = int(office_df["_O1"].sum())
        _aff_write_row(
            ws,
            row_i,
            col_offset,
            ["", f"{office} Total", "", "", "", office_leads, office_ftd, _cr(office_leads, office_ftd)],
            AFF_FONT_BOLD,
            AFF_FILL_OFFICE,
        )
        row_i += 1

    grand_leads = int(data_df["_N1"].sum())
    grand_ftd = int(data_df["_O1"].sum())
    _aff_write_row(
        ws,
        row_i,
        col_offset,
        ["GCC EN Total", "", "", "", "", grand_leads, grand_ftd, _cr(grand_leads, grand_ftd)],
        AFF_FONT_BOLD_W,
        AFF_FILL_GRAND,
    )


def build_aff_by_status(df, output_path, campaign_col, country_col, desk_col, status_col, n_col, o_col) -> None:
    data = df.copy()
    data["_OFFICE"] = data[desk_col].apply(get_office)
    data["_DESK2"] = data[desk_col].apply(get_desk2)
    data["_N1"] = data[n_col].apply(lambda x: 1 if str(x).strip() == "1" else 0)
    data["_O1"] = data[o_col].apply(lambda x: 1 if str(x).strip() == "1" else 0)

    ch_df = data[data[country_col].apply(lambda x: str(x).strip() == "Switzerland")].copy()
    sg_df = data[data[country_col].apply(lambda x: str(x).strip() == "Singapore")].copy()
    gcc_df = data[
        data[country_col].apply(lambda x: str(x).strip() in GCC_COUNTRIES) & (data["_DESK2"] == "EN")
    ].copy()

    wb = Workbook()
    ws = wb.active
    ws.title = "AFF by Status"

    _write_standard_table(
        ws,
        ch_df,
        col_offset=0,
        country_label="CH",
        campaign_col=campaign_col,
        status_col=status_col,
    )
    _write_standard_table(
        ws,
        sg_df,
        col_offset=8,
        country_label="SG",
        campaign_col=campaign_col,
        status_col=status_col,
    )
    _write_gcc_table(
        ws,
        gcc_df,
        col_offset=16,
        campaign_col=campaign_col,
        country_col=country_col,
        status_col=status_col,
    )

    col_widths = [
        10,
        9,
        22,
        18,
        8,
        7,
        7,
        1.5,
        10,
        9,
        22,
        18,
        8,
        7,
        7,
        1.5,
        10,
        9,
        22,
        20,
        18,
        8,
        7,
        7,
    ]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A2"
    wb.save(output_path)


def build_outputs(
    input_path: Path,
    output_dir: Path,
    lead_output_name: str | None = None,
    aff_output_name: str | None = None,
) -> list[Path]:
    today = datetime.now()
    output_dir.mkdir(parents=True, exist_ok=True)

    lead_name = lead_output_name or f"Lead Splitter - {today.strftime('%d-%m')}.xlsx"
    lead_output_path = output_dir / lead_name

    df = pd.read_excel(input_path, header=2, dtype=str)
    df.columns = [str(c) for c in df.columns]
    cols = df.columns.tolist()

    def col(idx):
        return cols[idx] if idx < len(cols) else None

    b_col = col(1)
    c_col = col(2)
    e_col = col(4)
    f_col = col(5)
    i_col = col(8)
    n_col = col(13)
    o_col = col(14)

    campaign_col = next((c for c in df.columns if str(c).strip().lower() == "campaign"), None)

    def both_empty(row):
        n_val = str(row[n_col]).strip() if pd.notna(row[n_col]) else ""
        o_val = str(row[o_col]).strip() if pd.notna(row[o_col]) else ""
        return n_val == "" and o_val == ""

    df = df[~df.apply(both_empty, axis=1)].copy()

    cid_counts = df[e_col].value_counts()
    dup_cids = cid_counts[cid_counts > 1].index.tolist()
    rows_to_drop = []

    for cid in dup_cids:
        group = df[df[e_col] == cid]
        for idx, row in group.iterrows():
            n1 = str(row[n_col]).strip() == "1" if pd.notna(row[n_col]) else False
            o1 = str(row[o_col]).strip() == "1" if pd.notna(row[o_col]) else False
            if n1 and not o1:
                rows_to_drop.append(idx)
            elif o1:
                df.at[idx, n_col] = "1"

    df = df.drop(index=rows_to_drop).copy()
    df = df[~df[c_col].apply(is_pool_agent)].copy()
    df[c_col] = df[c_col].apply(clean_agent_name).apply(fix_name)

    def update_desk(row):
        country = str(row[i_col]).strip() if pd.notna(row[i_col]) else ""
        agent = str(row[c_col]).strip() if pd.notna(row[c_col]) else ""
        if country == "Malaysia" and agent not in MY_EXEMPT_AGENTS:
            return "TR1-MY"
        if country == "Bangladesh":
            return "TR1-IN"
        return row[b_col]

    df[b_col] = df.apply(update_desk, axis=1)

    wb_orig = load_workbook(input_path, data_only=True)
    ws_orig = wb_orig.active
    header_row = [cell.value for cell in ws_orig[3]]

    wb_new = Workbook()
    ws_data = wb_new.active
    ws_data.title = "Data"
    ws_data.append(header_row)

    hdr_fill = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws_data[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = hdr_align

    data_border = make_border()
    row_font = Font(name="Arial", size=10)
    alt_fill = PatternFill("solid", start_color="EBF3FB", end_color="EBF3FB")

    for i, (_, row) in enumerate(df.iterrows()):
        row_data = [row[c] if pd.notna(row[c]) else None for c in df.columns]
        ws_data.append(row_data)
        excel_row = i + 2
        fill = alt_fill if i % 2 == 0 else PatternFill(fill_type=None)
        for cell in ws_data[excel_row]:
            cell.font = row_font
            cell.border = data_border
            cell.alignment = Alignment(vertical="center")
            cell.fill = fill

    for col_idx, col_cells in enumerate(ws_data.columns, 1):
        max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
        ws_data.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    ws_data.freeze_panes = "A2"
    build_pivot(wb_new, df, n_col, o_col, b_col, c_col, i_col)
    wb_new.save(lead_output_path)

    outputs = [lead_output_path]

    if campaign_col is not None:
        aff_name = aff_output_name or f"AFF BY status- SG - CH - GCC - {today.strftime('%d-%m')}.xlsx"
        aff_output_path = output_dir / aff_name
        build_aff_by_status(df, aff_output_path, campaign_col, i_col, b_col, f_col, n_col, o_col)
        outputs.append(aff_output_path)

    return outputs


def main() -> None:
    script_dir = Path(os.path.dirname(os.path.abspath(__file__)))
    input_path = script_dir / "report.xlsx"
    outputs = build_outputs(input_path=input_path, output_dir=script_dir)
    print("\n".join([f"Generated: {path.name}" for path in outputs]))


if __name__ == "__main__":
    main()
