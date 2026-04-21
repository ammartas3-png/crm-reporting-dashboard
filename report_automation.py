#!/usr/bin/env python3
"""
CRM + PowerBI report automation utility.

This script:
1) enriches CRM records with customer numbers from a PowerBI export,
2) optionally enriches extra columns (e.g. purple columns) from a 3rd file,
3) supports exact and fuzzy comment matching,
4) produces lead call counts, AFF/Status ratios, and call-frequency tables.
"""

from __future__ import annotations

import argparse
import re
import unicodedata
from pathlib import Path
from typing import Any

import pandas as pd
import yaml
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from rapidfuzz import fuzz, process

EMAIL_PATTERN = re.compile(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", re.IGNORECASE)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Automate CRM report generation using CRM and PowerBI files."
    )
    parser.add_argument("--crm", required=True, help="Path to CRM export (csv/xlsx).")
    parser.add_argument(
        "--powerbi", required=True, help="Path to PowerBI export (csv/xlsx)."
    )
    parser.add_argument(
        "--purple",
        help=(
            "Optional path to extra source file (csv/xlsx) used for columns such as "
            "Comments/Call Att."
        ),
    )
    parser.add_argument(
        "--config",
        required=True,
        help="Path to YAML config with column mappings and match settings.",
    )
    parser.add_argument(
        "--output",
        default="report_output.xlsx",
        help="Output Excel path. Default: report_output.xlsx",
    )
    return parser.parse_args()


def read_table(path: str) -> pd.DataFrame:
    file_path = Path(path)
    suffix = file_path.suffix.lower()

    if suffix == ".csv":
        return pd.read_csv(file_path)
    if suffix in {".xlsx", ".xls"}:
        return pd.read_excel(file_path)

    raise ValueError(
        f"Unsupported file type: {suffix}. Use csv/xlsx/xls for {file_path.name}"
    )


def normalize_text(value: Any) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip().lower()
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^a-z0-9\s]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def normalize_join_key(value: Any) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip().casefold()


def extract_comment_payload(comment_entry: str) -> str:
    parts = [part.strip() for part in str(comment_entry).split("|")]
    if len(parts) >= 3:
        return "|".join(parts[2:]).strip()
    if len(parts) == 2:
        return parts[1].strip()
    return parts[0].strip() if parts else ""


def should_skip_comment(comment_text: str, ignore_keywords: list[str]) -> bool:
    if not comment_text:
        return True

    lowered = comment_text.casefold()
    if EMAIL_PATTERN.search(comment_text):
        return True
    return any(keyword.casefold() in lowered for keyword in ignore_keywords)


def transform_comment_history(raw_value: Any, options: dict[str, Any] | None = None) -> Any:
    if pd.isna(raw_value):
        return pd.NA

    options = options or {}
    reverse_order = bool(options.get("reverse_order", True))
    output_separator = str(options.get("output_separator", " ; "))
    ignore_keywords = options.get("ignore_keywords", ["email"])
    if not isinstance(ignore_keywords, list):
        ignore_keywords = ["email"]

    text = str(raw_value)
    # Source comments can be separated by line breaks and/or semicolons.
    chunks = re.split(r"(?:\r?\n)+|;", text)
    parsed_comments: list[str] = []
    for chunk in chunks:
        clean_chunk = chunk.strip()
        if not clean_chunk:
            continue

        payload = extract_comment_payload(clean_chunk).strip()
        if should_skip_comment(payload, ignore_keywords):
            continue
        parsed_comments.append(payload)

    if reverse_order:
        parsed_comments = list(reversed(parsed_comments))
    if not parsed_comments:
        return pd.NA
    return output_separator.join(parsed_comments)


def apply_mapping_transform(source_values: pd.Series, mapping: dict[str, Any]) -> pd.Series:
    transform_name = mapping.get("transform")
    if not transform_name:
        return source_values

    options = mapping.get("transform_options", {})
    if transform_name == "comment_history":
        return source_values.map(lambda value: transform_comment_history(value, options=options))

    raise ValueError(
        f"Unsupported transform '{transform_name}' in purple_source.columns mapping for "
        f"target '{mapping.get('target', '')}'"
    )


def load_config(path: str) -> dict[str, Any]:
    with open(path, "r", encoding="utf-8") as file:
        config = yaml.safe_load(file) or {}
    return config


def validate_config(config: dict[str, Any]) -> None:
    required_paths = [
        ("columns", dict),
        ("columns.crm", dict),
        ("columns.powerbi", dict),
    ]
    for key_path, expected_type in required_paths:
        node = config
        for part in key_path.split("."):
            if part not in node:
                raise ValueError(f"Missing config section: '{key_path}'")
            node = node[part]
        if not isinstance(node, expected_type):
            raise ValueError(f"Config section '{key_path}' must be {expected_type.__name__}")

    required_crm_cols = ["lead", "customer_no", "comment", "aff", "status"]
    required_power_cols = ["lead", "customer_no", "comment"]

    for col_key in required_crm_cols:
        if col_key not in config["columns"]["crm"]:
            raise ValueError(f"Missing config key: columns.crm.{col_key}")
    for col_key in required_power_cols:
        if col_key not in config["columns"]["powerbi"]:
            raise ValueError(f"Missing config key: columns.powerbi.{col_key}")

    if "purple_source" in config:
        purple_cfg = config["purple_source"]
        if not isinstance(purple_cfg, dict):
            raise ValueError("Config section 'purple_source' must be dict")

        stage = purple_cfg.get("stage", "before_customer_match")
        if stage not in {"before_customer_match", "after_customer_match"}:
            raise ValueError(
                "purple_source.stage must be 'before_customer_match' or "
                "'after_customer_match'"
            )

        if "join" not in purple_cfg or not isinstance(purple_cfg["join"], dict):
            raise ValueError("Missing or invalid config section: purple_source.join")
        if "crm_key" not in purple_cfg["join"] or "source_key" not in purple_cfg["join"]:
            raise ValueError(
                "Missing config keys: purple_source.join.crm_key/source_key"
            )

        if "columns" not in purple_cfg or not isinstance(purple_cfg["columns"], list):
            raise ValueError("Missing or invalid config section: purple_source.columns")
        if not purple_cfg["columns"]:
            raise ValueError("Config section 'purple_source.columns' cannot be empty")

        for idx, col_map in enumerate(purple_cfg["columns"]):
            if not isinstance(col_map, dict):
                raise ValueError(
                    f"purple_source.columns[{idx}] must be an object "
                    "(source/target/overwrite)"
                )
            if "source" not in col_map or "target" not in col_map:
                raise ValueError(
                    f"Missing config keys in purple_source.columns[{idx}]: source/target"
                )
            transform_name = col_map.get("transform")
            if transform_name and transform_name not in {"comment_history"}:
                raise ValueError(
                    f"Unsupported transform in purple_source.columns[{idx}]: {transform_name}"
                )
            if (
                "transform_options" in col_map
                and not isinstance(col_map["transform_options"], dict)
            ):
                raise ValueError(
                    f"purple_source.columns[{idx}].transform_options must be an object"
                )


def validate_required_columns(df: pd.DataFrame, required_cols: list[str], label: str) -> None:
    missing = [col for col in required_cols if col not in df.columns]
    if missing:
        raise ValueError(f"{label} missing required columns: {missing}")


def find_best_match(
    crm_comment_norm: str,
    candidates: pd.DataFrame,
    power_customer_col: str,
    fuzzy_threshold: int,
) -> tuple[Any, float, str]:
    candidates = candidates[
        candidates["comment_norm"].astype(bool) & candidates[power_customer_col].notna()
    ]
    if candidates.empty:
        return None, 0.0, "not_found"

    exact = candidates[candidates["comment_norm"] == crm_comment_norm]
    if not exact.empty:
        customer_no = exact.iloc[0][power_customer_col]
        return customer_no, 100.0, "exact_comment"

    choice_map = {idx: text for idx, text in candidates["comment_norm"].items()}
    match = process.extractOne(
        crm_comment_norm,
        choice_map,
        scorer=fuzz.token_set_ratio,
        score_cutoff=fuzzy_threshold,
    )
    if not match:
        return None, 0.0, "not_found"

    _, score, matched_idx = match
    customer_no = candidates.loc[matched_idx, power_customer_col]
    return customer_no, float(score), "fuzzy_comment"


def enrich_crm_with_customer_numbers(
    crm_df: pd.DataFrame,
    power_df: pd.DataFrame,
    config: dict[str, Any],
) -> tuple[pd.DataFrame, pd.DataFrame]:
    crm_cols = config["columns"]["crm"]
    power_cols = config["columns"]["powerbi"]
    match_cfg = config.get("match", {})

    fuzzy_threshold = int(match_cfg.get("fuzzy_threshold", 80))
    lead_gate = bool(match_cfg.get("lead_gate", True))

    crm_lead_col = crm_cols["lead"]
    crm_customer_col = crm_cols["customer_no"]
    crm_comment_col = crm_cols["comment"]
    power_lead_col = power_cols["lead"]
    power_customer_col = power_cols["customer_no"]
    power_comment_col = power_cols["comment"]

    validate_required_columns(
        crm_df,
        [crm_lead_col, crm_customer_col, crm_comment_col],
        "CRM file",
    )
    validate_required_columns(
        power_df,
        [power_lead_col, power_customer_col, power_comment_col],
        "PowerBI file",
    )

    crm = crm_df.copy()
    power = power_df.copy()

    crm["lead_norm"] = crm[crm_lead_col].map(normalize_text)
    crm["comment_norm"] = crm[crm_comment_col].map(normalize_text)
    power["lead_norm"] = power[power_lead_col].map(normalize_text)
    power["comment_norm"] = power[power_comment_col].map(normalize_text)

    power_by_lead: dict[str, pd.DataFrame] = {
        lead: group.copy() for lead, group in power.groupby("lead_norm", dropna=False)
    }

    match_rows: list[dict[str, Any]] = []
    matched_customers: list[Any] = []
    match_scores: list[float] = []
    match_methods: list[str] = []

    for _, crm_row in crm.iterrows():
        current_customer = crm_row[crm_customer_col]
        if pd.notna(current_customer) and str(current_customer).strip():
            matched_customers.append(current_customer)
            match_scores.append(100.0)
            match_methods.append("already_present")
            match_rows.append(
                {
                    "crm_lead": crm_row[crm_lead_col],
                    "crm_comment": crm_row[crm_comment_col],
                    "resolved_customer_no": current_customer,
                    "score": 100.0,
                    "method": "already_present",
                }
            )
            continue

        crm_comment_norm = crm_row["comment_norm"]
        crm_lead_norm = crm_row["lead_norm"]
        if not crm_comment_norm:
            matched_customers.append(None)
            match_scores.append(0.0)
            match_methods.append("empty_comment")
            match_rows.append(
                {
                    "crm_lead": crm_row[crm_lead_col],
                    "crm_comment": crm_row[crm_comment_col],
                    "resolved_customer_no": None,
                    "score": 0.0,
                    "method": "empty_comment",
                }
            )
            continue

        if lead_gate and crm_lead_norm in power_by_lead:
            candidates = power_by_lead[crm_lead_norm]
        else:
            candidates = power

        customer_no, score, method = find_best_match(
            crm_comment_norm=crm_comment_norm,
            candidates=candidates,
            power_customer_col=power_customer_col,
            fuzzy_threshold=fuzzy_threshold,
        )
        matched_customers.append(customer_no)
        match_scores.append(score)
        match_methods.append(method)

        match_rows.append(
            {
                "crm_lead": crm_row[crm_lead_col],
                "crm_comment": crm_row[crm_comment_col],
                "resolved_customer_no": customer_no,
                "score": score,
                "method": method,
            }
        )

    crm["resolved_customer_no"] = matched_customers
    crm["match_score"] = match_scores
    crm["match_method"] = match_methods
    crm["customer_no_original"] = crm[crm_customer_col]
    crm[crm_customer_col] = crm[crm_customer_col].where(
        crm[crm_customer_col].notna(), crm["resolved_customer_no"]
    )

    match_audit = pd.DataFrame(match_rows)
    return crm, match_audit


def enrich_crm_with_purple_columns(
    crm_df: pd.DataFrame, purple_df: pd.DataFrame, config: dict[str, Any]
) -> tuple[pd.DataFrame, pd.DataFrame]:
    purple_cfg = config.get("purple_source")
    if not purple_cfg:
        return crm_df.copy(), pd.DataFrame()

    join_cfg = purple_cfg["join"]
    mappings = purple_cfg["columns"]
    crm_key = join_cfg["crm_key"]
    source_key = join_cfg["source_key"]
    source_cols = [m["source"] for m in mappings]

    validate_required_columns(crm_df, [crm_key], "CRM file")
    validate_required_columns(purple_df, [source_key, *source_cols], "Purple source file")

    crm = crm_df.copy()
    source = purple_df.copy()

    crm["_join_key"] = crm[crm_key].map(normalize_join_key)
    source["_join_key"] = source[source_key].map(normalize_join_key)
    source = source[source["_join_key"].astype(bool)]

    source_rows_before_dedupe = len(source)
    source = source.drop_duplicates(subset="_join_key", keep="first")
    duplicate_source_rows = source_rows_before_dedupe - len(source)

    source_renamed_cols: dict[str, str] = {}
    for idx, mapping in enumerate(mappings):
        source_renamed_cols[mapping["source"]] = f"_purple_{idx}"
    source = source.rename(columns=source_renamed_cols)

    merge_columns = ["_join_key", *source_renamed_cols.values()]
    merged = crm.merge(source[merge_columns], on="_join_key", how="left")

    audit_rows: list[dict[str, Any]] = []
    for idx, mapping in enumerate(mappings):
        target_col = mapping["target"]
        source_col = source_renamed_cols[mapping["source"]]
        overwrite = bool(mapping.get("overwrite", True))

        source_values = apply_mapping_transform(merged[source_col], mapping)
        old_values = merged[target_col].copy() if target_col in merged.columns else pd.Series(
            [pd.NA] * len(merged), index=merged.index
        )

        if overwrite:
            new_values = source_values.combine_first(old_values)
        else:
            new_values = old_values.combine_first(source_values)

        equal_or_both_na = old_values.eq(new_values) | (old_values.isna() & new_values.isna())
        merged[target_col] = new_values

        audit_rows.append(
            {
                "target_column": target_col,
                "source_column": mapping["source"],
                "overwrite": overwrite,
                "transform": mapping.get("transform", ""),
                "source_non_null_rows": int(source_values.notna().sum()),
                "applied_rows": int((~equal_or_both_na).sum()),
            }
        )

    matched_rows = int(merged["_join_key"].isin(set(source["_join_key"])).sum())
    unmatched_rows = int(len(merged) - matched_rows)

    audit_rows.extend(
        [
            {
                "target_column": "__meta__",
                "source_column": "crm_rows",
                "overwrite": None,
                "source_non_null_rows": None,
                "applied_rows": int(len(merged)),
            },
            {
                "target_column": "__meta__",
                "source_column": "matched_rows",
                "overwrite": None,
                "source_non_null_rows": None,
                "applied_rows": matched_rows,
            },
            {
                "target_column": "__meta__",
                "source_column": "unmatched_rows",
                "overwrite": None,
                "source_non_null_rows": None,
                "applied_rows": unmatched_rows,
            },
            {
                "target_column": "__meta__",
                "source_column": "duplicate_source_rows_skipped",
                "overwrite": None,
                "source_non_null_rows": None,
                "applied_rows": int(duplicate_source_rows),
            },
        ]
    )

    merged = merged.drop(columns=["_join_key", *source_renamed_cols.values()])
    return merged, pd.DataFrame(audit_rows)


def build_summary_tables(crm_enriched: pd.DataFrame, config: dict[str, Any]) -> dict[str, pd.DataFrame]:
    crm_cols = config["columns"]["crm"]
    lead_col = crm_cols["lead"]
    aff_col = crm_cols["aff"]
    status_col = crm_cols["status"]

    validate_required_columns(
        crm_enriched,
        [lead_col, aff_col, status_col],
        "Enriched CRM",
    )

    lead_call_counts = (
        crm_enriched.groupby(lead_col, dropna=False)
        .size()
        .reset_index(name="call_count")
        .sort_values("call_count", ascending=False)
    )

    call_count_distribution = (
        lead_call_counts["call_count"]
        .value_counts()
        .rename_axis("call_count")
        .reset_index(name="lead_count")
        .sort_values("call_count")
    )
    total_leads = call_count_distribution["lead_count"].sum()
    call_count_distribution["ratio_overall"] = (
        (call_count_distribution["lead_count"] / total_leads).round(4)
        if total_leads
        else 0.0
    )

    aff_status_counts = (
        crm_enriched.groupby([aff_col, status_col], dropna=False)
        .size()
        .reset_index(name="call_count")
    )
    aff_totals = aff_status_counts.groupby(aff_col)["call_count"].transform("sum")
    aff_status_counts["ratio_within_aff"] = (
        aff_status_counts["call_count"] / aff_totals
    ).round(4)
    aff_status_counts = aff_status_counts.sort_values(
        [aff_col, "call_count"], ascending=[True, False]
    )

    overall_status_summary = (
        crm_enriched.groupby(status_col, dropna=False)
        .size()
        .reset_index(name="call_count")
        .sort_values("call_count", ascending=False)
    )
    total_calls = overall_status_summary["call_count"].sum()
    overall_status_summary["ratio_overall"] = (
        (overall_status_summary["call_count"] / total_calls).round(4)
        if total_calls
        else 0.0
    )

    lead_aff_status_calls = (
        crm_enriched.groupby([lead_col, aff_col, status_col], dropna=False)
        .size()
        .reset_index(name="call_count")
    )

    call_frequency_by_aff_status = (
        lead_aff_status_calls.groupby([aff_col, status_col, "call_count"], dropna=False)
        .size()
        .reset_index(name="lead_count")
        .sort_values([aff_col, status_col, "call_count"])
    )

    return {
        "lead_call_counts": lead_call_counts,
        "call_count_distribution": call_count_distribution,
        "aff_status_ratios": aff_status_counts,
        "overall_status_summary": overall_status_summary,
        "call_frequency_by_aff_status": call_frequency_by_aff_status,
    }


def build_aff_status_dashboard_rows(
    aff_status_ratios: pd.DataFrame, aff_col: str, status_col: str
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for aff_value, group in aff_status_ratios.groupby(aff_col, dropna=False):
        aff_label = "BLANK" if pd.isna(aff_value) else str(aff_value)
        group_total = int(group["call_count"].sum())
        for _, row in group.iterrows():
            status_label = "BLANK" if pd.isna(row[status_col]) else str(row[status_col])
            rows.append(
                {
                    "aff": aff_label,
                    "status": status_label,
                    "count": int(row["call_count"]),
                    "ratio": float(row["ratio_within_aff"]),
                    "is_total": False,
                }
            )
        rows.append(
            {
                "aff": aff_label,
                "status": "TOTAL",
                "count": group_total,
                "ratio": 1.0,
                "is_total": True,
            }
        )
    rows.append(
        {
            "aff": "GRAND TOTAL",
            "status": "",
            "count": int(aff_status_ratios["call_count"].sum()),
            "ratio": 1.0,
            "is_total": True,
        }
    )
    return rows


def write_dashboard_sheet(workbook: Any, summaries: dict[str, pd.DataFrame], config: dict[str, Any]) -> None:
    ws = workbook.create_sheet("dashboard_summary")
    crm_cols = config["columns"]["crm"]
    aff_col = crm_cols["aff"]
    status_col = crm_cols["status"]

    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    total_fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def write_table_block(
        start_row: int,
        start_col: int,
        title: str,
        headers: list[str],
        rows: list[dict[str, Any]],
        percent_key: str,
    ) -> None:
        title_cell = ws.cell(row=start_row, column=start_col, value=title)
        title_cell.font = Font(bold=True)
        title_cell.alignment = Alignment(horizontal="center")
        ws.merge_cells(
            start_row=start_row,
            start_column=start_col,
            end_row=start_row,
            end_column=start_col + len(headers) - 1,
        )
        for col_idx, header in enumerate(headers):
            cell = ws.cell(row=start_row + 1, column=start_col + col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        data_start = start_row + 2
        data_end = data_start + len(rows) - 1

        for row_offset, row in enumerate(rows):
            current_row = data_start + row_offset
            values = [row["c1"], row["c2"], row["c3"], row["c4"]]
            for col_offset, value in enumerate(values):
                cell = ws.cell(row=current_row, column=start_col + col_offset, value=value)
                cell.border = thin_border
                if col_offset == 3:
                    cell.number_format = "0.00%"
                    cell.alignment = Alignment(horizontal="center")
                elif col_offset == 2:
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.alignment = Alignment(horizontal="left")
                if row["is_total"]:
                    cell.fill = total_fill
                    cell.font = Font(bold=True)

        if rows:
            percent_col_letter = ws.cell(row=data_start, column=start_col + 3).column_letter
            ws.conditional_formatting.add(
                f"{percent_col_letter}{data_start}:{percent_col_letter}{data_end}",
                ColorScaleRule(
                    start_type="num",
                    start_value=0,
                    start_color="F8696B",
                    mid_type="num",
                    mid_value=0.5,
                    mid_color="FFEB84",
                    end_type="num",
                    end_value=1,
                    end_color="63BE7B",
                ),
            )

    aff_rows = build_aff_status_dashboard_rows(
        summaries["aff_status_ratios"], aff_col=aff_col, status_col=status_col
    )
    aff_block = [
        {
            "c1": row["aff"],
            "c2": row["status"],
            "c3": row["count"],
            "c4": row["ratio"],
            "is_total": row["is_total"],
        }
        for row in aff_rows
    ]
    write_table_block(
        start_row=2,
        start_col=1,
        title="AFF + Status Breakdown",
        headers=["AFF", "Status", "Count", "Ratio"],
        rows=aff_block,
        percent_key="c4",
    )

    overall_rows = [
        {
            "c1": ("BLANK" if pd.isna(row[status_col]) else str(row[status_col])),
            "c2": "",
            "c3": int(row["call_count"]),
            "c4": float(row["ratio_overall"]),
            "is_total": False,
        }
        for _, row in summaries["overall_status_summary"].iterrows()
    ]
    overall_rows.append(
        {
            "c1": "TOTAL",
            "c2": "",
            "c3": int(summaries["overall_status_summary"]["call_count"].sum()),
            "c4": 1.0,
            "is_total": True,
        }
    )
    write_table_block(
        start_row=2,
        start_col=7,
        title="Overall Status Distribution",
        headers=["Status", "", "Count", "Ratio"],
        rows=overall_rows,
        percent_key="c4",
    )

    call_dist_rows = [
        {
            "c1": int(row["call_count"]),
            "c2": "",
            "c3": int(row["lead_count"]),
            "c4": float(row["ratio_overall"]),
            "is_total": False,
        }
        for _, row in summaries["call_count_distribution"].iterrows()
    ]
    call_dist_rows.append(
        {
            "c1": "TOTAL",
            "c2": "",
            "c3": int(summaries["call_count_distribution"]["lead_count"].sum()),
            "c4": 1.0,
            "is_total": True,
        }
    )
    write_table_block(
        start_row=14,
        start_col=7,
        title="Lead Call Count Distribution",
        headers=["Call Count", "", "Lead Count", "Ratio"],
        rows=call_dist_rows,
        percent_key="c4",
    )

    column_widths = {
        1: 18,
        2: 24,
        3: 10,
        4: 12,
        7: 18,
        8: 2,
        9: 12,
        10: 12,
    }
    for col_index, width in column_widths.items():
        ws.column_dimensions[ws.cell(row=1, column=col_index).column_letter].width = width


def write_output_excel(
    output_path: str,
    crm_enriched: pd.DataFrame,
    match_audit: pd.DataFrame,
    purple_merge_audit: pd.DataFrame,
    summaries: dict[str, pd.DataFrame],
    config: dict[str, Any],
) -> None:
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        crm_enriched.to_excel(writer, sheet_name="crm_enriched", index=False)
        match_audit.to_excel(writer, sheet_name="match_audit", index=False)
        if not purple_merge_audit.empty:
            purple_merge_audit.to_excel(writer, sheet_name="purple_merge_audit", index=False)
        for sheet_name in [
            "lead_call_counts",
            "call_count_distribution",
            "aff_status_ratios",
            "overall_status_summary",
            "call_frequency_by_aff_status",
        ]:
            df = summaries[sheet_name]
            df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
        write_dashboard_sheet(writer.book, summaries=summaries, config=config)


def main() -> None:
    args = parse_args()
    config = load_config(args.config)
    validate_config(config)

    crm_df = read_table(args.crm)
    power_df = read_table(args.powerbi)
    purple_merge_audit = pd.DataFrame()
    purple_cfg = config.get("purple_source", {})
    purple_stage = purple_cfg.get("stage", "before_customer_match")

    if args.purple and purple_stage == "before_customer_match":
        if "purple_source" not in config:
            raise ValueError(
                "--purple verildi fakat config içinde 'purple_source' bölümü bulunamadı."
            )
        purple_df = read_table(args.purple)
        crm_df, purple_merge_audit = enrich_crm_with_purple_columns(
            crm_df=crm_df, purple_df=purple_df, config=config
        )

    crm_enriched, match_audit = enrich_crm_with_customer_numbers(
        crm_df=crm_df,
        power_df=power_df,
        config=config,
    )

    if args.purple and purple_stage == "after_customer_match":
        if "purple_source" not in config:
            raise ValueError(
                "--purple verildi fakat config içinde 'purple_source' bölümü bulunamadı."
            )
        purple_df = read_table(args.purple)
        crm_enriched, purple_merge_audit = enrich_crm_with_purple_columns(
            crm_df=crm_enriched, purple_df=purple_df, config=config
        )

    summaries = build_summary_tables(crm_enriched=crm_enriched, config=config)
    lead_counts_map = dict(
        zip(
            summaries["lead_call_counts"][config["columns"]["crm"]["lead"]],
            summaries["lead_call_counts"]["call_count"],
        )
    )
    crm_enriched["lead_call_count"] = crm_enriched[config["columns"]["crm"]["lead"]].map(lead_counts_map)

    write_output_excel(
        output_path=args.output,
        crm_enriched=crm_enriched,
        match_audit=match_audit,
        purple_merge_audit=purple_merge_audit,
        summaries=summaries,
        config=config,
    )

    resolved_count = crm_enriched["resolved_customer_no"].notna().sum()
    total_rows = len(crm_enriched)
    print(
        f"Done. Output saved to {args.output} | "
        f"Customer number resolved rows: {resolved_count}/{total_rows}"
    )


if __name__ == "__main__":
    main()
