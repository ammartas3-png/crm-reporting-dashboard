from __future__ import annotations

import base64
import json
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from api import generate as api_generate
import program_b_country_report
from report_generator import (
    CRM_COLUMNS,
    OUTPUT_COLUMNS,
    POWERBI_COLUMNS,
    STATUS_LIST,
    build_output,
    read_powerbi_lookup,
)


def _write_workbook(path: Path, headers: list[str], rows: list[list[object]]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    workbook.save(path)


def _write_powerbi_with_third_row_headers(
    path: Path,
    rows: list[list[object]],
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["Applied filters: example"])
    worksheet.append([])
    worksheet.append(POWERBI_COLUMNS)
    for row in rows:
        worksheet.append(row)
    workbook.save(path)


class ReportGeneratorTests(unittest.TestCase):
    def test_build_output_merges_comments_and_call_attempts(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            powerbi = root / "powerbi.xlsx"
            crm = root / "crm.xlsx"
            output = root / "output.xlsx"

            _write_workbook(
                powerbi,
                POWERBI_COLUMNS,
                [
                    [
                        123,
                        "BrandA",
                        "| L1 called client ; | NA ; | VM ; | email follow up ;",
                        3,
                    ],
                ],
            )
            _write_workbook(
                crm,
                CRM_COLUMNS,
                [
                    [
                        "Lead",
                        123,
                        "2026-05-09",
                        "Jane Doe",
                        "Sales",
                        "Potential",
                        "TR",
                        "Campaign A",
                        "Sub A",
                        "Placement A",
                        "Agent 1",
                    ],
                    [
                        "Depositor",
                        456,
                        "2026-05-09",
                        "John Doe",
                        "Sales",
                        "Potential",
                        "TR",
                        "Campaign B",
                        "Sub B",
                        "Placement B",
                        "Agent 2",
                    ],
                ],
            )

            build_output(
                powerbi_report=powerbi,
                crm_files=[crm],
                platforms=["BrandA"],
                pivot_name="Status Pivot",
                output_file=output,
            )

            workbook = load_workbook(output, data_only=False)
            worksheet = workbook["CRM Output"]
            headers = [cell.value for cell in worksheet[1]]
            self.assertEqual(headers, OUTPUT_COLUMNS)

            comments_col = OUTPUT_COLUMNS.index("Comments") + 1
            attempts_col = OUTPUT_COLUMNS.index("Call Attempts") + 1
            status_col = OUTPUT_COLUMNS.index("Status") + 1

            self.assertEqual(worksheet.cell(2, comments_col).value, "NA VM x2 // called client")
            self.assertEqual(worksheet.cell(2, attempts_col).value, 3)
            self.assertEqual(worksheet.cell(3, status_col).value, "Telemarketing")
            self.assertEqual(worksheet.cell(3, attempts_col).value, 1)

    def test_missing_powerbi_columns_reports_file_name(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "bad_powerbi.xlsx"
            _write_workbook(path, ["Account No"], [[1]])

            with self.assertRaisesRegex(ValueError, "bad_powerbi.xlsx is missing required columns"):
                read_powerbi_lookup(path)

    def test_powerbi_headers_can_be_on_third_row(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "filtered_powerbi.xlsx"
            _write_powerbi_with_third_row_headers(
                path,
                [[123, "BrandA", "| L1 reached ;", 2]],
            )

            lookup = read_powerbi_lookup(path)

            self.assertEqual(lookup[("123", "branda")]["Call Attempts"], 2)
            self.assertEqual(lookup[("123", "branda")]["Comments"], ["L1 reached"])


class VercelBlobApiTests(unittest.TestCase):
    def test_blob_client_token_contains_upload_scope(self) -> None:
        with patch.dict(
            "os.environ",
            {"BLOB_READ_WRITE_TOKEN": "vercel_blob_rw_store123_secret"},
        ):
            token = api_generate._generate_blob_client_token("inputs/example.xlsx")

        self.assertTrue(token.startswith("vercel_blob_client_store123_"))
        encoded_token = token.split("_", 4)[4]
        signed_payload = base64.b64decode(encoded_token).decode("utf-8")
        _, encoded_payload = signed_payload.split(".", 1)
        payload = json.loads(base64.b64decode(encoded_payload).decode("utf-8"))

        self.assertEqual(payload["pathname"], "inputs/example.xlsx")
        self.assertEqual(payload["access"], "public")
        self.assertTrue(payload["addRandomSuffix"])
        self.assertIn(api_generate.XLSX_CONTENT_TYPE, payload["allowedContentTypes"])

    def test_blob_url_validation_rejects_non_blob_hosts(self) -> None:
        self.assertTrue(
            api_generate._is_blob_url("https://abc.public.blob.vercel-storage.com/a.xlsx")
        )
        self.assertFalse(api_generate._is_blob_url("https://example.com/a.xlsx"))


class ProgramBCountryReportTests(unittest.TestCase):
    def test_powerbi_headers_can_be_on_third_row(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "filtered_powerbi.xlsx"
            _write_powerbi_with_third_row_headers(
                path,
                [[123, "BrandA", "| L1 reached ;", 0]],
            )

            lookup = program_b_country_report.read_powerbi_lookup(path)

            self.assertEqual(lookup[("123", "branda")]["Call Attempts"], 1)
            self.assertEqual(lookup[("123", "branda")]["Comments"], ["L1 reached"])

    def test_build_output_creates_main_and_country_sheets(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            powerbi = root / "powerbi.xlsx"
            crm = root / "crm.xlsx"
            output = root / "country_output.xlsx"

            _write_workbook(
                powerbi,
                POWERBI_COLUMNS,
                [
                    [123, "BrandA", "| NA ; | VM ;", 0],
                    [456, "BrandA", "| L1 interested ;", 5],
                ],
            )
            _write_workbook(
                crm,
                CRM_COLUMNS,
                [
                    [
                        "Lead",
                        123,
                        "2026-05-09",
                        "Jane Doe",
                        "Sales",
                        "Potential",
                        "TR",
                        "Campaign A",
                        "Sub A",
                        "Placement A",
                        "Agent 1",
                    ],
                    [
                        "Lead",
                        456,
                        "2026-05-09",
                        "Max Doe",
                        "Sales",
                        "Call Again",
                        "DE",
                        "Campaign B",
                        "Sub B",
                        "Placement B",
                        "Agent 2",
                    ],
                ],
            )

            program_b_country_report.build_output(
                powerbi_report=powerbi,
                crm_files=[crm],
                platforms=["BrandA"],
                output_file=output,
            )

            workbook = load_workbook(output, data_only=False)
            self.assertIn("Main Report", workbook.sheetnames)
            self.assertIn("DE", workbook.sheetnames)
            self.assertIn("TR", workbook.sheetnames)

            main = workbook["Main Report"]
            headers = [cell.value for cell in main[1]]
            self.assertEqual(headers, OUTPUT_COLUMNS)

            attempts_col = OUTPUT_COLUMNS.index("Call Attempts") + 1
            self.assertEqual(main.cell(3, attempts_col).value, 1)
            self.assertEqual(main.cell(2, attempts_col).value, 5)

            status_pivot_start_row = 6
            status_pivot_status_col = 7
            status_pivot_count_col = 8
            status_pivot_pct_col = 9
            pivot_statuses = [
                main.cell(status_pivot_start_row + index, status_pivot_status_col).value
                for index in range(len(STATUS_LIST))
            ]
            self.assertEqual(pivot_statuses, STATUS_LIST)
            self.assertEqual(
                main.cell(status_pivot_start_row + 1, status_pivot_count_col).value,
                '=COUNTIF(\'Main Report\'!$G$2:$G$3,"Decline")',
            )
            self.assertEqual(
                main.cell(status_pivot_start_row + 1, status_pivot_pct_col).value,
                "=IFERROR(H7/H23,0)",
            )

            de_sheet = workbook["DE"]
            self.assertEqual([cell.value for cell in de_sheet[1]], OUTPUT_COLUMNS)
            self.assertTrue(str(de_sheet.cell(2, 1).value).startswith("=IFERROR("))
            self.assertEqual(de_sheet.cell(5, 6).value, "DE")
            self.assertEqual(de_sheet.cell(5, 7).value, "='Main Report'!$G$6")


if __name__ == "__main__":
    unittest.main()
